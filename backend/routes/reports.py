from fastapi import APIRouter, HTTPException, Depends, Response
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone, timedelta
import io

from core.database import db
from core.auth import get_user

router = APIRouter()


def _parse_report_dates(report_type, date, now):
    if report_type == "weekly":
        start = datetime.strptime(date, "%Y-%m-%d") if date else now - timedelta(days=now.weekday())
        start_str = start.strftime("%Y-%m-%d")
        end_str = (start + timedelta(days=6)).strftime("%Y-%m-%d")
        prev_start = (start - timedelta(days=7)).strftime("%Y-%m-%d")
        prev_end = (start - timedelta(days=1)).strftime("%Y-%m-%d")
    elif report_type == "monthly":
        start = datetime.strptime(date[:7] + "-01", "%Y-%m-%d") if date else now.replace(day=1)
        start_str = start.strftime("%Y-%m-%d")
        end = (start.replace(month=start.month + 1, day=1) if start.month < 12 else start.replace(year=start.year + 1, month=1, day=1)) - timedelta(days=1)
        end_str = end.strftime("%Y-%m-%d")
        prev_month = start - timedelta(days=1)
        prev_start = prev_month.replace(day=1).strftime("%Y-%m-%d")
        prev_end = prev_month.strftime("%Y-%m-%d")
    elif report_type == "quarterly":
        if date:
            parts = date.replace("Q", "").replace("q", "").split("-")
            year = int(parts[0])
            quarter = int(parts[1]) if len(parts) > 1 else ((now.month - 1) // 3 + 1)
        else:
            year = now.year
            quarter = (now.month - 1) // 3 + 1
        q_month = (quarter - 1) * 3 + 1
        start_str = f"{year}-{q_month:02d}-01"
        end_month = q_month + 2
        end_date = datetime(year, end_month, 1) + timedelta(days=31)
        end_date = end_date.replace(day=1) - timedelta(days=1)
        end_str = end_date.strftime("%Y-%m-%d")
        prev_q = quarter - 1 if quarter > 1 else 4
        prev_year = year if quarter > 1 else year - 1
        prev_q_month = (prev_q - 1) * 3 + 1
        prev_start = f"{prev_year}-{prev_q_month:02d}-01"
        prev_end_month = prev_q_month + 2
        prev_end_date = datetime(prev_year, prev_end_month, 1) + timedelta(days=31)
        prev_end_date = prev_end_date.replace(day=1) - timedelta(days=1)
        prev_end = prev_end_date.strftime("%Y-%m-%d")
    else:  # yearly
        year = int(date) if date else now.year
        start_str, end_str = f"{year}-01-01", f"{year}-12-31"
        prev_start, prev_end = f"{year-1}-01-01", f"{year-1}-12-31"
    return start_str, end_str, prev_start, prev_end


async def _build_report(rid, report_type, date):
    now = datetime.now(timezone.utc)
    start_str, end_str, prev_start, prev_end = _parse_report_dates(report_type, date, now)

    _appr = {"$or": [{"approval_status": {"$exists": False}}, {"approval_status": "approved"}]}

    purchases = await db.purchases.find({"restaurant_id": rid, **_appr, "invoice_date": {"$gte": start_str, "$lte": end_str}}, {"_id": 0}).to_list(10000)
    sales = await db.sales.find({"restaurant_id": rid, **_appr, "report_date": {"$gte": start_str, "$lte": end_str}}, {"_id": 0}).to_list(10000)
    salaries_cur = await db.salaries.find({"restaurant_id": rid, **_appr, "payment_date": {"$gte": start_str, "$lte": end_str}}, {"_id": 0}).to_list(10000)
    other_exp_cur = await db.other_expenses.find({"restaurant_id": rid, **_appr, "expense_date": {"$gte": start_str, "$lte": end_str}}, {"_id": 0}).to_list(10000)

    prev_purchases = await db.purchases.find({"restaurant_id": rid, **_appr, "invoice_date": {"$gte": prev_start, "$lte": prev_end}}, {"_id": 0}).to_list(10000)
    prev_sales = await db.sales.find({"restaurant_id": rid, **_appr, "report_date": {"$gte": prev_start, "$lte": prev_end}}, {"_id": 0}).to_list(10000)
    salaries_prev = await db.salaries.find({"restaurant_id": rid, **_appr, "payment_date": {"$gte": prev_start, "$lte": prev_end}}, {"_id": 0}).to_list(10000)
    other_exp_prev = await db.other_expenses.find({"restaurant_id": rid, **_appr, "expense_date": {"$gte": prev_start, "$lte": prev_end}}, {"_id": 0}).to_list(10000)

    total_p = round(sum(p["total"] for p in purchases), 2)
    total_s = round(sum(s["total_sales"] for s in sales), 2)
    total_sal = round(sum(s["amount"] for s in salaries_cur), 2)
    total_oe = round(sum(e["amount"] for e in other_exp_cur), 2)
    total_expenses = round(total_p + total_sal + total_oe, 2)
    net_profit = round(total_s - total_expenses, 2)

    prev_p = round(sum(p["total"] for p in prev_purchases), 2)
    prev_s = round(sum(s["total_sales"] for s in prev_sales), 2)
    prev_sal = round(sum(s["amount"] for s in salaries_prev), 2)
    prev_oe = round(sum(e["amount"] for e in other_exp_prev), 2)
    prev_total_expenses = round(prev_p + prev_sal + prev_oe, 2)
    prev_net_profit = round(prev_s - prev_total_expenses, 2)

    sup_spend = {}
    sup_invoice_count = {}
    for p in purchases:
        n = p.get("supplier_name", "Unknown")
        sup_spend[n] = sup_spend.get(n, 0) + p["total"]
        sup_invoice_count[n] = sup_invoice_count.get(n, 0) + 1

    item_spend = {}
    for p in purchases:
        for it in p.get("items", []):
            n = it.get("raw_name", "Unknown")
            item_spend[n] = item_spend.get(n, 0) + float(it.get("total", 0))

    def item_prices(plist):
        prices = {}
        for p in plist:
            for it in p.get("items", []):
                n = it.get("raw_name", "Unknown")
                price = float(it.get("unit_price", 0))
                if price > 0:
                    prices.setdefault(n, []).append(price)
        return {n: round(sum(v)/len(v), 2) for n, v in prices.items()}

    cur_prices = item_prices(purchases)
    prev_prices = item_prices(prev_purchases)
    price_changes = []
    for name, cur_p in cur_prices.items():
        prev_p_val = prev_prices.get(name)
        if prev_p_val and prev_p_val > 0:
            pct = round(((cur_p - prev_p_val) / prev_p_val) * 100, 1)
            if abs(pct) > 0:
                price_changes.append({"item": name, "current_price": cur_p, "previous_price": prev_p_val, "change_pct": pct})
    price_changes.sort(key=lambda x: -abs(x["change_pct"]))

    daily = {}
    for p in purchases:
        d = p.get("invoice_date", "")
        daily.setdefault(d, {"date": d, "purchases": 0, "sales": 0})
        daily[d]["purchases"] += p["total"]
    for s in sales:
        d = s.get("report_date", "")
        daily.setdefault(d, {"date": d, "purchases": 0, "sales": 0})
        daily[d]["sales"] += s["total_sales"]

    alerts = await db.alerts.find({"restaurant_id": rid}, {"_id": 0}).to_list(100)

    return {
        "report_type": report_type, "date_range": {"start": start_str, "end": end_str},
        "prev_date_range": {"start": prev_start, "end": prev_end},
        "total_purchases": total_p, "total_sales": total_s, "profit": round(total_s - total_p, 2),
        "prev_purchases": prev_p, "prev_sales": prev_s, "prev_profit": round(prev_s - prev_p, 2),
        "margin_pct": round((total_s - total_p) / total_s * 100, 1) if total_s > 0 else 0,
        "total_salaries": total_sal, "total_other_expenses": total_oe,
        "total_expenses": total_expenses, "net_profit": net_profit,
        "prev_salaries": prev_sal, "prev_other_expenses": prev_oe,
        "prev_total_expenses": prev_total_expenses, "prev_net_profit": prev_net_profit,
        "net_margin_pct": round(net_profit / total_s * 100, 1) if total_s > 0 else 0,
        "spending_by_supplier": [{"name": n, "total": round(t, 2), "invoices": sup_invoice_count.get(n, 0)} for n, t in sorted(sup_spend.items(), key=lambda x: -x[1])],
        "top_items": [{"name": n, "total": round(t, 2)} for n, t in sorted(item_spend.items(), key=lambda x: -x[1])[:10]],
        "price_changes": price_changes[:20],
        "daily_breakdown": sorted(daily.values(), key=lambda x: x["date"]),
        "alerts": alerts, "purchase_count": len(purchases), "sales_count": len(sales)
    }


@router.get("/reports")
async def get_reports(user=Depends(get_user), report_type: str = "weekly", date: str = ""):
    return await _build_report(user["restaurant_id"], report_type, date)


@router.get("/reports/category/{category}")
async def get_category_report(category: str, user=Depends(get_user), date_from: str = "", date_to: str = "", vendor: str = ""):
    """Get detailed report for a specific category with from/to date filtering."""
    rid = user["restaurant_id"]
    now = datetime.now(timezone.utc)
    if not date_from:
        date_from = now.strftime("%Y-%m-01")
    if not date_to:
        date_to = now.strftime("%Y-%m-%d")
    _appr = {"$or": [{"approval_status": {"$exists": False}}, {"approval_status": "approved"}]}

    if category == "sales":
        sales = await db.sales.find(
            {"restaurant_id": rid, **_appr, "$or": [
                {"report_date": {"$gte": date_from, "$lte": date_to}},
                {"date_from": {"$gte": date_from, "$lte": date_to}},
            ]}, {"_id": 0}
        ).sort("report_date", -1).to_list(5000)
        total = round(sum(s.get("total_sales", 0) for s in sales), 2)
        avg_per_entry = round(total / len(sales), 2) if sales else 0
        return {"category": "sales", "date_from": date_from, "date_to": date_to,
                "total_sales": total, "record_count": len(sales), "avg_per_entry": avg_per_entry, "records": sales}

    elif category == "raw_materials":
        purchases = await db.purchases.find(
            {"restaurant_id": rid, **_appr, "invoice_date": {"$gte": date_from, "$lte": date_to}}, {"_id": 0}
        ).sort("invoice_date", -1).to_list(5000)
        total = round(sum(p.get("total", 0) for p in purchases), 2)
        all_items = []
        for p in purchases:
            for it in p.get("items", []):
                all_items.append({
                    "vendor": p.get("supplier_name", ""), "date": p.get("invoice_date", ""),
                    "invoice": p.get("invoice_number", ""), "item": it.get("raw_name", ""),
                    "quantity": it.get("quantity", 0), "unit": it.get("unit", ""),
                    "unit_price": it.get("unit_price", 0), "line_total": it.get("total", 0)
                })
        return {"category": "raw_materials", "date_from": date_from, "date_to": date_to,
                "total": total, "invoice_count": len(purchases), "items": all_items, "records": purchases}

    elif category == "salaries":
        salaries = await db.salaries.find(
            {"restaurant_id": rid, **_appr, "payment_date": {"$gte": date_from, "$lte": date_to}}, {"_id": 0}
        ).sort("payment_date", -1).to_list(5000)
        total = round(sum(s.get("amount", 0) for s in salaries), 2)
        return {"category": "salaries", "date_from": date_from, "date_to": date_to,
                "total": total, "record_count": len(salaries), "records": salaries}

    elif category == "other_expenses":
        expenses = await db.other_expenses.find(
            {"restaurant_id": rid, **_appr, "expense_date": {"$gte": date_from, "$lte": date_to}}, {"_id": 0}
        ).sort("expense_date", -1).to_list(5000)
        total = round(sum(e.get("amount", 0) for e in expenses), 2)
        by_cat = {}
        for e in expenses:
            c = e.get("category", "Other")
            by_cat[c] = by_cat.get(c, 0) + e.get("amount", 0)
        breakdown = [{"category": k, "total": round(v, 2)} for k, v in sorted(by_cat.items(), key=lambda x: -x[1])]
        return {"category": "other_expenses", "date_from": date_from, "date_to": date_to,
                "total": total, "record_count": len(expenses), "records": expenses, "breakdown": breakdown}

    elif category == "vendor":
        query = {"restaurant_id": rid, **_appr, "invoice_date": {"$gte": date_from, "$lte": date_to}}
        if vendor:
            query["supplier_name"] = {"$regex": f"^{vendor}$", "$options": "i"}
        purchases = await db.purchases.find(query, {"_id": 0}).sort("invoice_date", -1).to_list(5000)
        total = round(sum(p.get("total", 0) for p in purchases), 2)
        items = []
        for p in purchases:
            for it in p.get("items", []):
                items.append({
                    "vendor": p.get("supplier_name", ""), "date": p.get("invoice_date", ""),
                    "item": it.get("raw_name", ""), "quantity": it.get("quantity", 0),
                    "unit": it.get("unit", ""), "price": it.get("unit_price", 0),
                    "total": it.get("total", 0)
                })
        all_vendors = await db.suppliers.find({"restaurant_id": rid}, {"_id": 0, "name": 1}).to_list(200)
        vendor_names = sorted(set([v["name"] for v in all_vendors]))
        return {"category": "vendor", "date_from": date_from, "date_to": date_to, "vendor": vendor or "All",
                "total": total, "invoice_count": len(purchases), "items": items, "records": purchases, "vendors": vendor_names}

    elif category == "profit":
        sales = await db.sales.find(
            {"restaurant_id": rid, **_appr, "$or": [
                {"report_date": {"$gte": date_from, "$lte": date_to}},
                {"date_from": {"$gte": date_from, "$lte": date_to}},
            ]}, {"_id": 0}
        ).to_list(5000)
        purchases = await db.purchases.find(
            {"restaurant_id": rid, **_appr, "invoice_date": {"$gte": date_from, "$lte": date_to}}, {"_id": 0}
        ).to_list(5000)
        salaries = await db.salaries.find(
            {"restaurant_id": rid, **_appr, "payment_date": {"$gte": date_from, "$lte": date_to}}, {"_id": 0}
        ).to_list(5000)
        other_exp = await db.other_expenses.find(
            {"restaurant_id": rid, **_appr, "expense_date": {"$gte": date_from, "$lte": date_to}}, {"_id": 0}
        ).to_list(5000)
        total_sales = round(sum(s.get("total_sales", 0) for s in sales), 2)
        raw_mat = round(sum(p.get("total", 0) for p in purchases), 2)
        sal = round(sum(s.get("amount", 0) for s in salaries), 2)
        oe = round(sum(e.get("amount", 0) for e in other_exp), 2)
        total_exp = round(raw_mat + sal + oe, 2)
        net_profit = round(total_sales - total_exp, 2)
        margin = round(net_profit / total_sales * 100, 1) if total_sales > 0 else 0
        return {"category": "profit", "date_from": date_from, "date_to": date_to,
                "total_sales": total_sales, "raw_materials": raw_mat, "salaries": sal,
                "other_expenses": oe, "total_expenses": total_exp, "net_profit": net_profit, "net_margin_pct": margin}

    raise HTTPException(400, f"Unknown category: {category}")


@router.get("/reports/category/{category}/export")
async def export_category_report(category: str, fmt: str = "excel", user=Depends(get_user), date_from: str = "", date_to: str = "", vendor: str = ""):
    """Export a category report as PDF or Excel."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    report = await get_category_report(category, user, date_from, date_to, vendor)
    df = report.get("date_from", "")
    dt = report.get("date_to", "")
    title_map = {"sales": "Sales Report", "raw_materials": "Raw Material Expense Report", "salaries": "Salary Report",
                 "other_expenses": "Other Expense Report", "vendor": "Vendor Purchase Report", "profit": "Profit Report"}
    title = title_map.get(category, "Report")

    if fmt == "pdf":
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#0f172a'), spaceAfter=4*mm)
        sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#94a3b8'), spaceAfter=6*mm)
        elements = [Paragraph(title, title_style), Paragraph(f"{df} to {dt}", sub_style)]

        if category == "sales":
            data = [["Date", "Total Sales"]]
            for r in report.get("records", []):
                data.append([r.get("report_date", r.get("date_from", "")), f"${r.get('total_sales',0):,.2f}"])
            data.append(["TOTAL", f"${report['total_sales']:,.2f}"])
        elif category == "raw_materials":
            data = [["Vendor", "Item", "Date", "Qty", "Price", "Total"]]
            for it in report.get("items", []):
                data.append([it["vendor"], it["item"], it["date"], str(it["quantity"]), f"${it['unit_price']:,.2f}", f"${it['line_total']:,.2f}"])
            data.append(["", "", "", "", "TOTAL", f"${report['total']:,.2f}"])
        elif category == "salaries":
            data = [["Employee", "Position", "Amount", "Date"]]
            for r in report.get("records", []):
                data.append([r.get("employee_name",""), r.get("position",""), f"${r.get('amount',0):,.2f}", r.get("payment_date","")])
            data.append(["", "", f"${report['total']:,.2f}", "TOTAL"])
        elif category == "other_expenses":
            data = [["Title", "Category", "Amount", "Date", "Notes"]]
            for r in report.get("records", []):
                data.append([r.get("title",""), r.get("category",""), f"${r.get('amount',0):,.2f}", r.get("expense_date",""), (r.get("notes","") or "")[:30]])
            data.append(["", "", f"${report['total']:,.2f}", "TOTAL", ""])
        elif category == "vendor":
            data = [["Vendor", "Item", "Date", "Qty", "Price", "Total"]]
            for it in report.get("items", []):
                data.append([it["vendor"], it["item"], it["date"], str(it["quantity"]), f"${it['price']:,.2f}", f"${it['total']:,.2f}"])
            data.append(["", "", "", "", "TOTAL", f"${report['total']:,.2f}"])
        elif category == "profit":
            data = [["Category", "Amount"],
                    ["Total Sales", f"${report['total_sales']:,.2f}"], ["", ""],
                    ["Expenses:", ""], ["  Raw Materials", f"${report['raw_materials']:,.2f}"],
                    ["  Salaries", f"${report['salaries']:,.2f}"], ["  Other Expenses", f"${report['other_expenses']:,.2f}"],
                    ["Total Expenses", f"${report['total_expenses']:,.2f}"], ["", ""],
                    ["Net Profit", f"${report['net_profit']:,.2f}"], ["Net Margin", f"{report['net_margin_pct']}%"]]
        else:
            data = [["No data"]]

        t = RLTable(data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0f172a')), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTSIZE', (0,0), (-1,-1), 8), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
            ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'), ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#f1f5f9')),
        ]))
        elements.append(t)
        doc.build(elements)
        buf.seek(0)
        return Response(content=buf.read(), media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="{category}_report_{df}_{dt}.pdf"'})

    else:  # excel
        wb = Workbook()
        ws = wb.active
        ws.title = title
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")

        ws.append([title])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append([f"Period: {df} to {dt}"])
        ws.append([])

        if category == "sales":
            headers = ["Date", "Total Sales"]
            ws.append(headers)
            for c in range(1, len(headers)+1): ws.cell(row=4, column=c).font = header_font; ws.cell(row=4, column=c).fill = header_fill
            for r in report.get("records", []):
                ws.append([r.get("report_date", r.get("date_from", "")), r.get("total_sales", 0)])
            ws.append(["TOTAL", report["total_sales"]])
        elif category == "raw_materials":
            headers = ["Vendor", "Item", "Date", "Quantity", "Unit Price", "Total"]
            ws.append(headers)
            for c in range(1, len(headers)+1): ws.cell(row=4, column=c).font = header_font; ws.cell(row=4, column=c).fill = header_fill
            for it in report.get("items", []):
                ws.append([it["vendor"], it["item"], it["date"], it["quantity"], it["unit_price"], it["line_total"]])
            ws.append(["", "", "", "", "TOTAL", report["total"]])
        elif category == "salaries":
            headers = ["Employee", "Position", "Amount", "Payment Date", "Notes"]
            ws.append(headers)
            for c in range(1, len(headers)+1): ws.cell(row=4, column=c).font = header_font; ws.cell(row=4, column=c).fill = header_fill
            for r in report.get("records", []):
                ws.append([r.get("employee_name",""), r.get("position",""), r.get("amount",0), r.get("payment_date",""), r.get("notes","")])
            ws.append(["", "", report["total"], "TOTAL", ""])
        elif category == "other_expenses":
            headers = ["Title", "Category", "Amount", "Date", "Notes"]
            ws.append(headers)
            for c in range(1, len(headers)+1): ws.cell(row=4, column=c).font = header_font; ws.cell(row=4, column=c).fill = header_fill
            for r in report.get("records", []):
                ws.append([r.get("title",""), r.get("category",""), r.get("amount",0), r.get("expense_date",""), r.get("notes","")])
            ws.append(["", "", report["total"], "TOTAL", ""])
        elif category == "vendor":
            headers = ["Vendor", "Item", "Date", "Quantity", "Price", "Total"]
            ws.append(headers)
            for c in range(1, len(headers)+1): ws.cell(row=4, column=c).font = header_font; ws.cell(row=4, column=c).fill = header_fill
            for it in report.get("items", []):
                ws.append([it["vendor"], it["item"], it["date"], it["quantity"], it["price"], it["total"]])
            ws.append(["", "", "", "", "TOTAL", report["total"]])
        elif category == "profit":
            headers = ["Category", "Amount"]
            ws.append(headers)
            for c in range(1, len(headers)+1): ws.cell(row=4, column=c).font = header_font; ws.cell(row=4, column=c).fill = header_fill
            ws.append(["Total Sales", report["total_sales"]])
            ws.append([])
            ws.append(["Raw Materials", report["raw_materials"]])
            ws.append(["Salaries", report["salaries"]])
            ws.append(["Other Expenses", report["other_expenses"]])
            ws.append(["Total Expenses", report["total_expenses"]])
            ws.append([])
            ws.append(["Net Profit", report["net_profit"]])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

        for col_cells in ws.columns:
            max_length = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_length + 4, 30)

        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(content=buf.read(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f'attachment; filename="{category}_report_{df}_{dt}.xlsx"'})


@router.get("/reports/download")
async def download_report(user=Depends(get_user), report_type: str = "weekly", date: str = "", fmt: str = "excel"):
    report = await _build_report(user["restaurant_id"], report_type, date)

    if fmt == "pdf":
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title2', parent=styles['Title'], fontSize=16, spaceAfter=6)
        sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=9, textColor=colors.grey)
        section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=12, spaceBefore=14, spaceAfter=6)

        elements = []
        elements.append(Paragraph("Restaurant Financial Report", title_style))
        elements.append(Paragraph(f"{report_type.title()} &bull; {report['date_range']['start']} to {report['date_range']['end']}", sub_style))
        elements.append(Spacer(1, 8*mm))

        tax_data = [
            ['Category', 'Amount'],
            ['Total Sales (Revenue)', f"${report['total_sales']:,.2f}"],
            ['', ''],
            ['Expenses Breakdown:', ''],
            ['  Raw Materials', f"${report['total_purchases']:,.2f}"],
            ['  Salaries', f"${report['total_salaries']:,.2f}"],
            ['  Other Expenses', f"${report['total_other_expenses']:,.2f}"],
            ['Total Expenses', f"${report['total_expenses']:,.2f}"],
            ['', ''],
            ['Net Profit', f"${report['net_profit']:,.2f}"],
            ['Net Margin', f"{report['net_margin_pct']}%"],
        ]
        t = Table(tax_data, colWidths=[100*mm, 60*mm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0f172a')), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTSIZE', (0,0), (-1,-1), 9), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
            ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('BACKGROUND', (0,3), (-1,3), colors.HexColor('#f8fafc')), ('FONTNAME', (0,3), (-1,3), 'Helvetica-Bold'),
            ('BACKGROUND', (0,7), (-1,7), colors.HexColor('#f1f5f9')), ('FONTNAME', (0,7), (-1,7), 'Helvetica-Bold'),
            ('BACKGROUND', (0,9), (-1,9), colors.HexColor('#ecfdf5') if report['net_profit'] >= 0 else colors.HexColor('#fef2f2')),
            ('FONTNAME', (0,9), (-1,9), 'Helvetica-Bold'), ('FONTNAME', (0,10), (-1,10), 'Helvetica-Bold'),
        ]))
        elements.append(Paragraph("Tax Summary", section_style))
        elements.append(t)
        elements.append(Spacer(1, 4*mm))

        kpi_data = [['Revenue', 'Purchases', 'Profit', 'Margin'],
                     [f"${report['total_sales']:,.2f}", f"${report['total_purchases']:,.2f}", f"${report['profit']:,.2f}", f"{report['margin_pct']}%"]]
        t = Table(kpi_data, colWidths=[45*mm]*4)
        t.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0f172a')), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTSIZE', (0,0), (-1,-1), 9), ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
            ('TOPPADDING', (0,0), (-1,-1), 6), ('BOTTOMPADDING', (0,0), (-1,-1), 6)]))
        elements.append(t)

        if report['spending_by_supplier']:
            elements.append(Paragraph("Spending by Supplier", section_style))
            sup_data = [['Supplier', 'Total', 'Invoices']] + [[s['name'], f"${s['total']:,.2f}", str(s['invoices'])] for s in report['spending_by_supplier'][:10]]
            t = Table(sup_data, colWidths=[80*mm, 45*mm, 30*mm])
            t.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')), ('FONTSIZE', (0,0), (-1,-1), 8),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')), ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4)]))
            elements.append(t)

        if report['price_changes']:
            elements.append(Paragraph("Price Changes", section_style))
            pc_data = [['Item', 'Previous', 'Current', 'Change']] + [[p['item'], f"${p['previous_price']:,.2f}", f"${p['current_price']:,.2f}", f"{p['change_pct']:+.1f}%"] for p in report['price_changes'][:15]]
            t = Table(pc_data, colWidths=[60*mm, 35*mm, 35*mm, 30*mm])
            t.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')), ('FONTSIZE', (0,0), (-1,-1), 8),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')), ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4)]))
            elements.append(t)

        doc.build(elements)
        buf.seek(0)
        filename = f"report_{report_type}_{report['date_range']['start']}.pdf"
        return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})

    else:  # excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        header_font = Font(bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")

        ws = wb.active
        ws.title = "Summary"
        ws.append(["Restaurant Financial Report"])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append([f"{report_type.title()} Report: {report['date_range']['start']} to {report['date_range']['end']}"])
        ws.append([])
        ws.append(["Metric", "Current Period", "Previous Period", "Change"])
        for col in range(1, 5):
            cell = ws.cell(row=4, column=col)
            cell.font = header_font
            cell.fill = header_fill
        def pct_chg(cur, prev):
            return f"{((cur - prev) / prev * 100):+.1f}%" if prev else "N/A"
        ws.append(["Revenue", report['total_sales'], report['prev_sales'], pct_chg(report['total_sales'], report['prev_sales'])])
        ws.append(["Purchases", report['total_purchases'], report['prev_purchases'], pct_chg(report['total_purchases'], report['prev_purchases'])])
        ws.append(["Profit", report['profit'], report['prev_profit'], pct_chg(report['profit'], report['prev_profit']) if report['prev_profit'] != 0 else "N/A"])
        ws.append(["Margin %", f"{report['margin_pct']}%", "", ""])
        ws.append([])
        ws.append(["--- TAX SUMMARY ---"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
        ws.append(["Total Sales (Revenue)", report['total_sales']])
        ws.append([])
        ws.append(["Expenses Breakdown:"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(["  Raw Materials", report['total_purchases']])
        ws.append(["  Salaries", report['total_salaries']])
        ws.append(["  Other Expenses", report['total_other_expenses']])
        ws.append(["Total Expenses", report['total_expenses']])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append([])
        ws.append(["Net Profit", report['net_profit']])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
        ws.cell(row=ws.max_row, column=2).font = Font(bold=True, size=11)
        ws.append(["Net Margin", f"{report['net_margin_pct']}%"])
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 14

        ws2 = wb.create_sheet("Suppliers")
        ws2.append(["Supplier", "Total Spent", "Invoices"])
        for col in range(1, 4):
            cell = ws2.cell(row=1, column=col); cell.font = header_font; cell.fill = header_fill
        for s in report['spending_by_supplier']:
            ws2.append([s['name'], s['total'], s['invoices']])
        ws2.column_dimensions['A'].width = 30; ws2.column_dimensions['B'].width = 15; ws2.column_dimensions['C'].width = 12

        ws3 = wb.create_sheet("Price Changes")
        ws3.append(["Item", "Previous Price", "Current Price", "Change %"])
        for col in range(1, 5):
            cell = ws3.cell(row=1, column=col); cell.font = header_font; cell.fill = header_fill
        for p in report['price_changes']:
            ws3.append([p['item'], p['previous_price'], p['current_price'], p['change_pct']])
        ws3.column_dimensions['A'].width = 25; ws3.column_dimensions['B'].width = 15; ws3.column_dimensions['C'].width = 15; ws3.column_dimensions['D'].width = 12

        ws4 = wb.create_sheet("Daily Breakdown")
        ws4.append(["Date", "Purchases", "Sales"])
        for col in range(1, 4):
            cell = ws4.cell(row=1, column=col); cell.font = header_font; cell.fill = header_fill
        for d in report['daily_breakdown']:
            ws4.append([d['date'], round(d['purchases'], 2), round(d['sales'], 2)])
        ws4.column_dimensions['A'].width = 15; ws4.column_dimensions['B'].width = 15; ws4.column_dimensions['C'].width = 15

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"report_{report_type}_{report['date_range']['start']}.xlsx"
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})
