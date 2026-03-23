from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse, Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os, logging, uuid, base64, json, re, io
from pathlib import Path
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
from jose import jwt, JWTError
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
SECRET_KEY = os.environ.get("JWT_SECRET", "fallback-secret")
ALGORITHM = "HS256"
security = HTTPBearer(auto_error=False)
LLM_KEY = os.environ.get("EMERGENT_LLM_KEY", "")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== MODELS ====================

class UserRegister(BaseModel):
    email: str
    password: str
    name: str
    restaurant_name: str

class UserLogin(BaseModel):
    email: str
    password: str

class PurchaseCreate(BaseModel):
    supplier_name: str
    supplier_id: Optional[str] = None
    invoice_number: str
    invoice_date: str
    items: List[Dict[str, Any]]
    subtotal: float
    tax: float
    total: float

class PurchaseUpdate(BaseModel):
    supplier_name: Optional[str] = None
    supplier_id: Optional[str] = None
    invoice_number: Optional[str] = None
    invoice_date: Optional[str] = None
    items: Optional[List[Dict[str, Any]]] = None
    subtotal: Optional[float] = None
    tax: Optional[float] = None
    total: Optional[float] = None

class SalesCreate(BaseModel):
    report_date: Optional[str] = None  # kept for backward compat
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    total_sales: float
    items: Optional[List[Dict[str, Any]]] = []

class SalesUpdate(BaseModel):
    report_date: Optional[str] = None
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    total_sales: Optional[float] = None
    items: Optional[List[Dict[str, Any]]] = None

class SupplierCreate(BaseModel):
    name: str
    contact_person: Optional[str] = ""
    phone: Optional[str] = ""
    email: Optional[str] = ""
    address: Optional[str] = ""

class CanonicalItemCreate(BaseModel):
    name: str
    category: Optional[str] = ""

class ItemAliasCreate(BaseModel):
    canonical_item_id: str
    alias_name: str

class ChatMessageIn(BaseModel):
    message: str

class SalaryCreate(BaseModel):
    employee_name: str
    position: Optional[str] = ""
    amount: float
    payment_date: str
    notes: Optional[str] = ""

class OtherExpenseCreate(BaseModel):
    title: str
    category: str
    amount: float
    expense_date: str
    notes: Optional[str] = ""

class SettingsUpdate(BaseModel):
    name: Optional[str] = None
    restaurant_name: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    currency: Optional[str] = None
    default_tax_rate: Optional[float] = None
    default_expense_category: Optional[str] = None
    alerts_enabled: Optional[bool] = None
    alert_price_increase: Optional[bool] = None
    alert_cheaper_vendor: Optional[bool] = None
    alert_not_ordered: Optional[bool] = None
    language: Optional[str] = None
    date_format: Optional[str] = None

# ==================== AUTH UTILITIES ====================

def hash_pw(pw):
    return pwd_context.hash(pw)

def verify_pw(pw, h):
    return pwd_context.verify(pw, h)

def make_token(uid):
    return jwt.encode({"user_id": uid, "exp": datetime.now(timezone.utc) + timedelta(days=7)}, SECRET_KEY, algorithm=ALGORITHM)

async def get_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    if not credentials:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        user = await db.users.find_one({"id": payload.get("user_id")}, {"_id": 0})
        if not user:
            raise HTTPException(401, "User not found")
        if user.get("status") == "inactive":
            raise HTTPException(403, "Account is deactivated")
        return user
    except JWTError:
        raise HTTPException(401, "Invalid token")

def require_manager(user):
    if user.get("role") != "manager":
        raise HTTPException(403, "Manager access required")

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register")
async def register(data: UserRegister):
    if await db.users.find_one({"email": data.email}):
        raise HTTPException(400, "Email already registered")
    rid = str(uuid.uuid4())
    await db.restaurants.insert_one({"id": rid, "name": data.restaurant_name, "address": "", "phone": "", "created_at": datetime.now(timezone.utc).isoformat()})
    uid = str(uuid.uuid4())
    await db.users.insert_one({"id": uid, "email": data.email, "password_hash": hash_pw(data.password), "name": data.name, "restaurant_id": rid, "role": "manager", "status": "active", "created_at": datetime.now(timezone.utc).isoformat()})
    return {"token": make_token(uid), "user": {"id": uid, "email": data.email, "name": data.name, "restaurant_id": rid, "restaurant_name": data.restaurant_name, "role": "manager"}}

@api_router.post("/auth/login")
async def login(data: UserLogin):
    u = await db.users.find_one({"email": data.email}, {"_id": 0})
    if not u or not verify_pw(data.password, u["password_hash"]):
        raise HTTPException(401, "Invalid credentials")
    if u.get("status") == "inactive":
        raise HTTPException(403, "Account is deactivated. Contact your manager.")
    r = await db.restaurants.find_one({"id": u["restaurant_id"]}, {"_id": 0})
    return {"token": make_token(u["id"]), "user": {"id": u["id"], "email": u["email"], "name": u["name"], "restaurant_id": u["restaurant_id"], "restaurant_name": r["name"] if r else "", "role": u.get("role", "staff")}}

@api_router.get("/auth/me")
async def me(user=Depends(get_user)):
    r = await db.restaurants.find_one({"id": user["restaurant_id"]}, {"_id": 0})
    return {"id": user["id"], "email": user["email"], "name": user["name"], "restaurant_id": user["restaurant_id"], "restaurant_name": r["name"] if r else "", "role": user.get("role", "staff")}

# ==================== USER MANAGEMENT ====================

VALID_ROLES = ["manager", "accountant", "cashier", "staff"]

ALL_PERMISSIONS = [
    "can_add_sales", "can_edit_sales", "can_delete_sales",
    "can_add_expenses", "can_edit_expenses", "can_delete_expenses",
    "can_upload_files", "can_view_reports", "can_export_reports",
    "can_view_records", "can_manage_vendors", "can_manage_items",
    "can_manage_users",
]

DEFAULT_PERMISSIONS = {
    "manager": {p: True for p in ALL_PERMISSIONS},
    "accountant": {
        "can_add_sales": True, "can_edit_sales": True, "can_delete_sales": False,
        "can_add_expenses": True, "can_edit_expenses": True, "can_delete_expenses": False,
        "can_upload_files": True, "can_view_reports": True, "can_export_reports": True,
        "can_view_records": True, "can_manage_vendors": True, "can_manage_items": True,
        "can_manage_users": False,
    },
    "cashier": {
        "can_add_sales": True, "can_edit_sales": False, "can_delete_sales": False,
        "can_add_expenses": False, "can_edit_expenses": False, "can_delete_expenses": False,
        "can_upload_files": False, "can_view_reports": False, "can_export_reports": False,
        "can_view_records": False, "can_manage_vendors": False, "can_manage_items": False,
        "can_manage_users": False,
    },
    "staff": {
        "can_add_sales": False, "can_edit_sales": False, "can_delete_sales": False,
        "can_add_expenses": False, "can_edit_expenses": False, "can_delete_expenses": False,
        "can_upload_files": True, "can_view_reports": False, "can_export_reports": False,
        "can_view_records": False, "can_manage_vendors": False, "can_manage_items": False,
        "can_manage_users": False,
    },
}

class UserCreate(BaseModel):
    name: str
    email: str
    password: str
    role: str = "staff"
    permissions: Optional[Dict[str, bool]] = None
    approval_rule: str = "pending_all"
    auto_approve_limit: Optional[float] = None

class UserUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    password: Optional[str] = None
    role: Optional[str] = None
    status: Optional[str] = None
    permissions: Optional[Dict[str, bool]] = None
    approval_rule: Optional[str] = None
    auto_approve_limit: Optional[float] = None

VALID_APPROVAL_RULES = ["auto_approve_all", "auto_approve_below", "pending_all"]

def _safe_user(u):
    """Return user dict without password_hash, with permissions and approval defaulted."""
    out = {k: v for k, v in u.items() if k != "password_hash"}
    if "permissions" not in out:
        out["permissions"] = DEFAULT_PERMISSIONS.get(out.get("role", "staff"), DEFAULT_PERMISSIONS["staff"])
    if "approval_rule" not in out:
        out["approval_rule"] = "auto_approve_all" if out.get("role") == "manager" else "pending_all"
    if "auto_approve_limit" not in out:
        out["auto_approve_limit"] = None
    return out

def _compute_approval_status(user, amount):
    """Determine approval_status for a new record based on user's approval rule."""
    role = user.get("role", "staff")
    if role == "manager":
        return "approved"
    rule = user.get("approval_rule", "pending_all")
    if rule == "auto_approve_all":
        return "approved"
    if rule == "auto_approve_below":
        limit = user.get("auto_approve_limit")
        if limit is not None and amount <= limit:
            return "approved"
        return "pending"
    return "pending"

@api_router.get("/users")
async def list_users(user=Depends(get_user)):
    require_manager(user)
    rid = user["restaurant_id"]
    users = await db.users.find({"restaurant_id": rid}, {"_id": 0}).to_list(500)
    return [_safe_user(u) for u in users]

@api_router.post("/users")
async def create_user(data: UserCreate, user=Depends(get_user)):
    require_manager(user)
    rid = user["restaurant_id"]
    if data.role not in VALID_ROLES:
        raise HTTPException(400, f"Invalid role. Must be one of: {', '.join(VALID_ROLES)}")
    if await db.users.find_one({"email": data.email}):
        raise HTTPException(400, "Email already in use")
    if len(data.password) < 6:
        raise HTTPException(400, "Password must be at least 6 characters")
    uid = str(uuid.uuid4())
    if data.permissions:
        perms = {}
        for p in ALL_PERMISSIONS:
            perms[p] = bool(data.permissions.get(p, False))
    else:
        perms = DEFAULT_PERMISSIONS.get(data.role, DEFAULT_PERMISSIONS["staff"])
    doc = {
        "id": uid,
        "email": data.email,
        "password_hash": hash_pw(data.password),
        "name": data.name,
        "restaurant_id": rid,
        "role": data.role,
        "status": "active",
        "permissions": perms,
        "approval_rule": data.approval_rule if data.approval_rule in VALID_APPROVAL_RULES else "pending_all",
        "auto_approve_limit": data.auto_approve_limit,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"],
    }
    await db.users.insert_one(doc)
    doc.pop("_id", None)
    return _safe_user(doc)

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, data: UserUpdate, user=Depends(get_user)):
    require_manager(user)
    rid = user["restaurant_id"]
    target = await db.users.find_one({"id": user_id, "restaurant_id": rid}, {"_id": 0})
    if not target:
        raise HTTPException(404, "User not found")
    updates = {}
    if data.name is not None:
        updates["name"] = data.name
    if data.email is not None:
        existing = await db.users.find_one({"email": data.email, "id": {"$ne": user_id}})
        if existing:
            raise HTTPException(400, "Email already in use")
        updates["email"] = data.email
    if data.password is not None:
        if len(data.password) < 6:
            raise HTTPException(400, "Password must be at least 6 characters")
        updates["password_hash"] = hash_pw(data.password)
    if data.role is not None:
        if data.role not in VALID_ROLES:
            raise HTTPException(400, f"Invalid role. Must be one of: {', '.join(VALID_ROLES)}")
        if user_id == user["id"] and data.role != "manager":
            raise HTTPException(400, "You cannot change your own role")
        updates["role"] = data.role
    if data.status is not None:
        if data.status not in ("active", "inactive"):
            raise HTTPException(400, "Status must be 'active' or 'inactive'")
        if user_id == user["id"] and data.status == "inactive":
            raise HTTPException(400, "You cannot deactivate yourself")
        updates["status"] = data.status
    if data.permissions is not None:
        # Validate all keys are valid permissions
        clean_perms = {}
        for p in ALL_PERMISSIONS:
            clean_perms[p] = bool(data.permissions.get(p, False))
        updates["permissions"] = clean_perms
    if data.approval_rule is not None:
        if data.approval_rule not in VALID_APPROVAL_RULES:
            raise HTTPException(400, f"Invalid approval_rule. Must be one of: {', '.join(VALID_APPROVAL_RULES)}")
        updates["approval_rule"] = data.approval_rule
    if data.auto_approve_limit is not None:
        updates["auto_approve_limit"] = data.auto_approve_limit
    if not updates:
        raise HTTPException(400, "No fields to update")
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.users.update_one({"id": user_id}, {"$set": updates})
    updated = await db.users.find_one({"id": user_id}, {"_id": 0})
    return _safe_user(updated)

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, user=Depends(get_user)):
    require_manager(user)
    rid = user["restaurant_id"]
    if user_id == user["id"]:
        raise HTTPException(400, "You cannot delete yourself")
    target = await db.users.find_one({"id": user_id, "restaurant_id": rid}, {"_id": 0})
    if not target:
        raise HTTPException(404, "User not found")
    await db.users.delete_one({"id": user_id, "restaurant_id": rid})
    return {"status": "deleted"}

@api_router.get("/users/permissions/defaults")
async def get_default_permissions(user=Depends(get_user)):
    """Get the default permission presets for each role."""
    require_manager(user)
    return DEFAULT_PERMISSIONS

@api_router.put("/users/{user_id}/permissions")
async def update_user_permissions(user_id: str, permissions: Dict[str, bool], user=Depends(get_user)):
    """Update a user's permissions directly."""
    require_manager(user)
    rid = user["restaurant_id"]
    target = await db.users.find_one({"id": user_id, "restaurant_id": rid}, {"_id": 0})
    if not target:
        raise HTTPException(404, "User not found")
    clean_perms = {}
    for p in ALL_PERMISSIONS:
        clean_perms[p] = bool(permissions.get(p, False))
    await db.users.update_one({"id": user_id}, {"$set": {"permissions": clean_perms, "updated_at": datetime.now(timezone.utc).isoformat()}})
    updated = await db.users.find_one({"id": user_id}, {"_id": 0})
    return _safe_user(updated)



# ==================== APPROVALS ====================

class ApprovalAction(BaseModel):
    action: str  # "approve" or "reject"
    reason: Optional[str] = None

@api_router.get("/approvals")
async def list_pending_records(
    user=Depends(get_user),
    record_type: str = "",
    status: str = "pending",
    created_by: str = "",
    date_from: str = "",
    date_to: str = "",
):
    """List records pending approval (Manager only)."""
    require_manager(user)
    rid = user["restaurant_id"]
    results = []

    collections = {
        "sale": ("sales", "total_sales", "report_date"),
        "purchase": ("purchases", "total", "invoice_date"),
        "salary": ("salaries", "amount", "payment_date"),
        "other_expense": ("other_expenses", "amount", "expense_date"),
    }

    types_to_check = [record_type] if record_type and record_type in collections else list(collections.keys())

    for rtype in types_to_check:
        coll_name, amt_field, date_field = collections[rtype]
        coll = db[coll_name]
        query = {"restaurant_id": rid}
        if status:
            query["approval_status"] = status
        if created_by:
            query["created_by_id"] = created_by
        if date_from:
            query.setdefault(date_field, {})["$gte"] = date_from
        if date_to:
            query.setdefault(date_field, {})["$lte"] = date_to

        docs = await coll.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
        for d in docs:
            results.append({
                "record_type": rtype,
                "record_id": d.get("id", ""),
                "date": d.get(date_field, ""),
                "amount": d.get(amt_field, 0),
                "created_by_id": d.get("created_by_id", ""),
                "created_by_name": d.get("created_by_name", "Unknown"),
                "approval_status": d.get("approval_status", "approved"),
                "rejection_reason": d.get("rejection_reason", ""),
                "notes": d.get("notes", d.get("transaction_notes", "")),
                "created_at": d.get("created_at", ""),
                # Type-specific fields
                "supplier_name": d.get("supplier_name", ""),
                "employee_name": d.get("employee_name", ""),
                "title": d.get("title", ""),
                "category": d.get("category", ""),
                "invoice_number": d.get("invoice_number", ""),
            })

    results.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return results


@api_router.get("/approvals/counts")
async def approval_counts(user=Depends(get_user)):
    """Get count of pending records per type."""
    require_manager(user)
    rid = user["restaurant_id"]
    counts = {}
    for rtype, (coll_name, _, _) in {"sale": ("sales", "", ""), "purchase": ("purchases", "", ""), "salary": ("salaries", "", ""), "other_expense": ("other_expenses", "", "")}.items():
        counts[rtype] = await db[coll_name].count_documents({"restaurant_id": rid, "approval_status": "pending"})
    counts["total"] = sum(counts.values())
    return counts


@api_router.put("/approvals/{record_type}/{record_id}")
async def process_approval(record_type: str, record_id: str, data: ApprovalAction, user=Depends(get_user)):
    """Approve or reject a record."""
    require_manager(user)
    rid = user["restaurant_id"]

    coll_map = {"sale": "sales", "purchase": "purchases", "salary": "salaries", "other_expense": "other_expenses"}
    coll_name = coll_map.get(record_type)
    if not coll_name:
        raise HTTPException(400, "Invalid record_type")

    if data.action not in ("approve", "reject"):
        raise HTTPException(400, "Action must be 'approve' or 'reject'")

    coll = db[coll_name]
    rec = await coll.find_one({"id": record_id, "restaurant_id": rid}, {"_id": 0})
    if not rec:
        raise HTTPException(404, "Record not found")

    updates = {
        "approval_status": "approved" if data.action == "approve" else "rejected",
        "approved_by_id": user["id"],
        "approved_by_name": user.get("name", ""),
        "approved_at": datetime.now(timezone.utc).isoformat(),
    }
    if data.action == "reject" and data.reason:
        updates["rejection_reason"] = data.reason

    await coll.update_one({"id": record_id}, {"$set": updates})
    updated = await coll.find_one({"id": record_id}, {"_id": 0})
    return updated



# ==================== SMART ALERTS ENGINE ====================

async def _generate_smart_alerts(rid):
    """Analyze real purchase history and generate actionable smart alerts."""
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")

    purchases = await db.purchases.find({"restaurant_id": rid}, {"_id": 0}).to_list(10000)
    if not purchases:
        return []

    alerts = []

    # --- 1. ITEMS NOT ORDERED FOR A LONG TIME ---
    # Items purchased before but not recently (cutoff: 10 days)
    recent_cutoff = (now - timedelta(days=10)).strftime("%Y-%m-%d")
    older_cutoff = (now - timedelta(days=90)).strftime("%Y-%m-%d")

    item_last_purchase = {}  # item_name -> {date, vendor, price}
    for p in sorted(purchases, key=lambda x: x.get("invoice_date", "")):
        d = p.get("invoice_date", "")
        for it in p.get("items", []):
            name = it.get("raw_name", "").strip()
            if not name:
                continue
            item_last_purchase[name] = {
                "date": d,
                "vendor": p.get("supplier_name", ""),
                "price": float(it.get("unit_price", 0)),
            }

    for item_name, info in item_last_purchase.items():
        if info["date"] < recent_cutoff and info["date"] >= older_cutoff:
            days_ago = (now - datetime.strptime(info["date"], "%Y-%m-%d").replace(tzinfo=timezone.utc)).days
            severity = "high" if days_ago > 21 else ("medium" if days_ago > 14 else "low")
            alerts.append({
                "type": "not_ordered",
                "severity": severity,
                "item_name": item_name,
                "vendor": info["vendor"],
                "days_since": days_ago,
                "last_date": info["date"],
                "last_price": info["price"],
            })

    # Sort by days_since descending, limit to 8
    not_ordered = [a for a in alerts if a["type"] == "not_ordered"]
    not_ordered.sort(key=lambda x: -x["days_since"])
    alerts = not_ordered[:8]

    # --- 2. PRICE INCREASES ---
    # Compare most recent price per item vs the previous price
    item_price_history = {}  # item_name -> [(date, price, vendor)]
    for p in sorted(purchases, key=lambda x: x.get("invoice_date", "")):
        d = p.get("invoice_date", "")
        vendor = p.get("supplier_name", "")
        for it in p.get("items", []):
            name = it.get("raw_name", "").strip()
            price = float(it.get("unit_price", 0))
            if name and price > 0:
                item_price_history.setdefault(name, []).append({"date": d, "price": price, "vendor": vendor})

    price_alerts = []
    for name, history in item_price_history.items():
        if len(history) < 2:
            continue
        latest = history[-1]
        previous = history[-2]
        if latest["price"] > previous["price"] and previous["price"] > 0:
            pct = ((latest["price"] - previous["price"]) / previous["price"]) * 100
            if pct > 3:  # Only alert on >3% increases
                severity = "high" if pct > 15 else ("medium" if pct > 8 else "low")
                price_alerts.append({
                    "type": "price_increase",
                    "severity": severity,
                    "item_name": name,
                    "vendor": latest["vendor"],
                    "old_price": previous["price"],
                    "new_price": latest["price"],
                    "change_pct": round(pct, 1),
                    "old_date": previous["date"],
                    "new_date": latest["date"],
                    "old_vendor": previous["vendor"],
                })

    price_alerts.sort(key=lambda x: -x["change_pct"])
    alerts += price_alerts[:8]

    # --- 3. CHEAPER VENDOR ALTERNATIVES ---
    # For each item, find if another vendor sells it for less
    item_vendor_prices = {}  # item_name -> {vendor: {price, date}}
    for p in sorted(purchases, key=lambda x: x.get("invoice_date", "")):
        d = p.get("invoice_date", "")
        vendor = p.get("supplier_name", "")
        for it in p.get("items", []):
            name = it.get("raw_name", "").strip()
            price = float(it.get("unit_price", 0))
            if name and price > 0:
                item_vendor_prices.setdefault(name, {})[vendor] = {"price": price, "date": d}

    cheaper_alerts = []
    for name, vendors in item_vendor_prices.items():
        if len(vendors) < 2:
            continue
        sorted_vendors = sorted(vendors.items(), key=lambda x: x[1]["price"])
        cheapest_vendor, cheapest_info = sorted_vendors[0]
        # Check if most recent purchase was NOT from the cheapest vendor
        last_purchase = item_price_history.get(name, [])
        if not last_purchase:
            continue
        last = last_purchase[-1]
        if last["vendor"] != cheapest_vendor and last["price"] > cheapest_info["price"]:
            savings_pct = ((last["price"] - cheapest_info["price"]) / last["price"]) * 100
            if savings_pct > 3:
                severity = "high" if savings_pct > 20 else ("medium" if savings_pct > 10 else "low")
                cheaper_alerts.append({
                    "type": "cheaper_vendor",
                    "severity": severity,
                    "item_name": name,
                    "vendor": last["vendor"],
                    "current_price": last["price"],
                    "cheaper_vendor": cheapest_vendor,
                    "cheaper_price": cheapest_info["price"],
                    "savings_pct": round(savings_pct, 1),
                    "last_cheap_date": cheapest_info["date"],
                })

    cheaper_alerts.sort(key=lambda x: -x["savings_pct"])
    alerts += cheaper_alerts[:8]

    # Final sort: high first, then medium, then low
    sev_order = {"high": 0, "medium": 1, "low": 2}
    alerts.sort(key=lambda a: sev_order.get(a["severity"], 9))

    return alerts

# ==================== DASHBOARD ====================

@api_router.get("/dashboard/summary")
async def dashboard_summary(user=Depends(get_user)):
    rid = user["restaurant_id"]
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")
    month_start = now.strftime("%Y-%m-01")
    prev_month = now.replace(day=1) - timedelta(days=1)
    prev_month_start = prev_month.strftime("%Y-%m-01")
    prev_month_end = prev_month.strftime("%Y-%m-%d")

    _approved = {"$or": [{"approval_status": {"$exists": False}}, {"approval_status": "approved"}]}
    purchases = await db.purchases.find({"restaurant_id": rid, **_approved}, {"_id": 0}).to_list(10000)
    salaries = await db.salaries.find({"restaurant_id": rid, **_approved}, {"_id": 0}).to_list(10000)
    other_exp = await db.other_expenses.find({"restaurant_id": rid, **_approved}, {"_id": 0}).to_list(10000)

    def sum_p(df, dt=None):
        return sum(p["total"] for p in purchases if p.get("invoice_date", "") >= df and (not dt or p.get("invoice_date", "") <= dt))
    def sum_sal(df, dt=None):
        return sum(s["amount"] for s in salaries if s.get("payment_date", "") >= df and (not dt or s.get("payment_date", "") <= dt))
    def sum_oe(df, dt=None):
        return sum(e["amount"] for e in other_exp if e.get("expense_date", "") >= df and (not dt or e.get("expense_date", "") <= dt))

    smart_alerts = await _generate_smart_alerts(rid)
    # Limit to top 5 most actionable insights (high severity first)
    severity_order = {"high": 0, "medium": 1, "low": 2}
    smart_alerts.sort(key=lambda a: severity_order.get(a.get("severity", "low"), 2))
    smart_alerts = smart_alerts[:5]

    return {
        "month_raw_materials": round(sum_p(month_start, today), 2),
        "month_salaries": round(sum_sal(month_start, today), 2),
        "month_other_expenses": round(sum_oe(month_start, today), 2),
        "prev_month_raw_materials": round(sum_p(prev_month_start, prev_month_end), 2),
        "prev_month_salaries": round(sum_sal(prev_month_start, prev_month_end), 2),
        "prev_month_other_expenses": round(sum_oe(prev_month_start, prev_month_end), 2),
        "smart_alerts": smart_alerts,
    }


@api_router.get("/dashboard/item-search")
async def dashboard_item_search(q: str = "", user=Depends(get_user)):
    """Search for an item across all purchases and return vendor/price comparison."""
    rid = user["restaurant_id"]
    if not q or len(q.strip()) < 2:
        return {"results": []}

    query_lower = q.strip().lower()
    _approved = {"$or": [{"approval_status": {"$exists": False}}, {"approval_status": "approved"}]}
    purchases = await db.purchases.find({"restaurant_id": rid, **_approved}, {"_id": 0}).to_list(10000)

    # Build per-item, per-vendor price data
    item_vendor_data = {}
    for p in purchases:
        vendor = p.get("supplier_name", "Unknown")
        inv_date = p.get("invoice_date", "")
        for it in p.get("items", []):
            raw_name = it.get("raw_name", "")
            if not raw_name or query_lower not in raw_name.lower():
                continue
            key = raw_name.lower()
            if key not in item_vendor_data:
                item_vendor_data[key] = {"name": raw_name, "vendors": {}}
            vd = item_vendor_data[key]["vendors"]
            if vendor not in vd:
                vd[vendor] = {"prices": [], "dates": [], "unit": it.get("unit", "")}
            unit_price = float(it.get("unit_price", 0))
            if unit_price > 0:
                vd[vendor]["prices"].append(unit_price)
                vd[vendor]["dates"].append(inv_date)

    results = []
    for key, item_data in item_vendor_data.items():
        vendors = []
        for vname, vinfo in item_data["vendors"].items():
            if not vinfo["prices"]:
                continue
            latest_idx = vinfo["dates"].index(max(vinfo["dates"])) if vinfo["dates"] else 0
            vendors.append({
                "vendor": vname,
                "latest_price": round(vinfo["prices"][latest_idx], 2),
                "avg_price": round(sum(vinfo["prices"]) / len(vinfo["prices"]), 2),
                "min_price": round(min(vinfo["prices"]), 2),
                "max_price": round(max(vinfo["prices"]), 2),
                "purchase_count": len(vinfo["prices"]),
                "last_date": max(vinfo["dates"]) if vinfo["dates"] else "",
                "unit": vinfo["unit"],
            })
        if not vendors:
            continue
        vendors.sort(key=lambda v: v["latest_price"])
        cheapest = vendors[0]
        results.append({
            "item_name": item_data["name"],
            "vendors": vendors,
            "cheapest_vendor": cheapest["vendor"],
            "cheapest_price": cheapest["latest_price"],
            "vendor_count": len(vendors),
        })

    results.sort(key=lambda r: r["item_name"].lower())
    return {"results": results}


@api_router.get("/dashboard/drill-down/{category}")
async def dashboard_drill_down(category: str, user=Depends(get_user)):
    """Drill-down data for a spending category."""
    rid = user["restaurant_id"]
    now = datetime.now(timezone.utc)
    month_start = now.strftime("%Y-%m-01")
    today = now.strftime("%Y-%m-%d")
    _approved = {"$or": [{"approval_status": {"$exists": False}}, {"approval_status": "approved"}]}

    if category == "raw_materials":
        purchases = await db.purchases.find(
            {"restaurant_id": rid, **_approved}, {"_id": 0}
        ).to_list(10000)
        month_purchases = [p for p in purchases if month_start <= p.get("invoice_date", "") <= today]

        item_map = {}
        for p in month_purchases:
            vendor = p.get("supplier_name", "Unknown")
            supplier_id = p.get("supplier_id", "")
            inv_date = p.get("invoice_date", "")
            for it in p.get("items", []):
                name = it.get("raw_name", "Unknown")
                key = name.lower()
                if key not in item_map:
                    item_map[key] = {"name": name, "total_spent": 0, "vendors": {}}
                total = float(it.get("total", 0))
                item_map[key]["total_spent"] += total
                vd = item_map[key]["vendors"]
                if vendor not in vd:
                    vd[vendor] = {"prices": [], "dates": [], "unit": it.get("unit", ""), "supplier_id": supplier_id}
                up = float(it.get("unit_price", 0))
                if up > 0:
                    vd[vendor]["prices"].append(up)
                    vd[vendor]["dates"].append(inv_date)

        items = []
        for key, im in item_map.items():
            vendors = []
            for vname, vi in im["vendors"].items():
                if not vi["prices"]:
                    continue
                latest_idx = vi["dates"].index(max(vi["dates"])) if vi["dates"] else 0
                vendors.append({
                    "vendor": vname,
                    "supplier_id": vi["supplier_id"],
                    "latest_price": round(vi["prices"][latest_idx], 2),
                    "avg_price": round(sum(vi["prices"]) / len(vi["prices"]), 2),
                    "min_price": round(min(vi["prices"]), 2),
                    "max_price": round(max(vi["prices"]), 2),
                    "purchase_count": len(vi["prices"]),
                    "last_date": max(vi["dates"]) if vi["dates"] else "",
                    "unit": vi["unit"],
                })
            vendors.sort(key=lambda v: v["latest_price"])
            cheapest = vendors[0]["vendor"] if vendors else ""
            items.append({
                "item_name": im["name"],
                "total_spent": round(im["total_spent"], 2),
                "vendors": vendors,
                "cheapest_vendor": cheapest,
                "vendor_count": len(vendors),
            })
        items.sort(key=lambda x: -x["total_spent"])
        return {"category": "raw_materials", "items": items, "total": round(sum(i["total_spent"] for i in items), 2)}

    elif category == "salaries":
        sals = await db.salaries.find(
            {"restaurant_id": rid, **_approved}, {"_id": 0}
        ).to_list(10000)
        month_sals = [s for s in sals if month_start <= s.get("payment_date", "") <= today]
        employees = []
        for s in month_sals:
            employees.append({
                "name": s.get("employee_name", "Unknown"),
                "position": s.get("position", ""),
                "amount": round(s.get("amount", 0), 2),
                "payment_date": s.get("payment_date", ""),
                "payment_method": s.get("payment_method", ""),
            })
        employees.sort(key=lambda e: -e["amount"])
        return {"category": "salaries", "employees": employees, "total": round(sum(e["amount"] for e in employees), 2)}

    elif category == "other":
        expenses = await db.other_expenses.find(
            {"restaurant_id": rid, **_approved}, {"_id": 0}
        ).to_list(10000)
        month_exp = [e for e in expenses if month_start <= e.get("expense_date", "") <= today]
        by_category = {}
        for e in month_exp:
            cat = e.get("category", "Uncategorized") or "Uncategorized"
            if cat not in by_category:
                by_category[cat] = {"items": [], "total": 0}
            by_category[cat]["items"].append({
                "title": e.get("title", "Untitled"),
                "amount": round(e.get("amount", 0), 2),
                "expense_date": e.get("expense_date", ""),
                "vendor": e.get("vendor", ""),
                "notes": e.get("notes", ""),
            })
            by_category[cat]["total"] += e.get("amount", 0)
        categories = []
        for cname, cdata in sorted(by_category.items(), key=lambda x: -x[1]["total"]):
            cdata["items"].sort(key=lambda x: -x["amount"])
            categories.append({
                "category_name": cname,
                "total": round(cdata["total"], 2),
                "items": cdata["items"],
            })
        return {"category": "other", "categories": categories, "total": round(sum(c["total"] for c in categories), 2)}

    return {"error": "Invalid category. Use: raw_materials, salaries, other"}

# ==================== UPLOAD / EXTRACT ====================

@api_router.post("/upload/parse-excel")
async def parse_excel(file: UploadFile = File(...), document_type: str = Form("purchase_invoice"), user=Depends(get_user)):
    """Parse Excel/CSV files and extract purchase or sales data."""
    import openpyxl, csv as csv_mod
    try:
        content = await file.read()
        fname = (file.filename or "").lower()
        rows = []

        if fname.endswith('.csv'):
            text = content.decode('utf-8', errors='replace')
            reader = csv_mod.reader(text.strip().splitlines())
            for r in reader:
                rows.append(r)
        elif fname.endswith('.xlsx') or fname.endswith('.xls'):
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            for r in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else '' for c in r])
            wb.close()
        else:
            raise HTTPException(400, "Unsupported file type. Use .xlsx, .xls, or .csv")

        if len(rows) < 2:
            raise HTTPException(400, "File has no data rows")

        # Normalize headers
        headers_raw = [str(h).strip().lower().replace(' ', '_') for h in rows[0]]
        col_map = {}
        for i, h in enumerate(headers_raw):
            for key, aliases in {
                'supplier': ['supplier', 'supplier_name', 'vendor', 'vendor_name', 'from'],
                'date': ['date', 'invoice_date', 'inv_date', 'purchase_date', 'order_date', 'report_date'],
                'invoice_number': ['invoice', 'invoice_number', 'inv_no', 'invoice_no', 'inv_number', 'invoice#', 'inv#', 'ref', 'reference'],
                'item_name': ['item', 'item_name', 'product', 'product_name', 'description', 'raw_name', 'name', 'menu_item', 'ingredient'],
                'quantity': ['quantity', 'qty', 'count'],
                'unit': ['unit', 'uom', 'measure', 'unit_of_measure'],
                'unit_price': ['price', 'unit_price', 'unit_cost', 'cost', 'rate'],
                'total': ['total', 'line_total', 'subtotal', 'ext_price', 'extended_price', 'revenue', 'amount'],
            }.items():
                if h in aliases and key not in col_map:
                    col_map[key] = i

        data_rows = rows[1:]

        def safe_float(val):
            try:
                s = str(val).replace('$', '').replace(',', '').strip()
                return float(s) if s else 0
            except (ValueError, TypeError):
                return 0

        def safe_date(val):
            s = str(val).strip()
            for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%m-%d-%Y', '%d-%m-%Y', '%Y/%m/%d']:
                try:
                    return datetime.strptime(s[:10], fmt).strftime('%Y-%m-%d')
                except (ValueError, TypeError):
                    continue
            return datetime.now(timezone.utc).strftime('%Y-%m-%d')

        if document_type == "purchase_invoice":
            # Group by supplier + date + invoice to create purchases
            items_parsed = []
            for row in data_rows:
                if len(row) <= max(col_map.values(), default=0):
                    row.extend([''] * (max(col_map.values(), default=0) + 1 - len(row)))
                item_name = row[col_map['item_name']].strip() if 'item_name' in col_map else ''
                if not item_name:
                    continue
                qty = safe_float(row[col_map['quantity']]) if 'quantity' in col_map else 1
                up = safe_float(row[col_map['unit_price']]) if 'unit_price' in col_map else 0
                tot = safe_float(row[col_map['total']]) if 'total' in col_map else (qty * up)
                if tot == 0 and qty > 0 and up > 0:
                    tot = qty * up
                if up == 0 and tot > 0 and qty > 0:
                    up = tot / qty

                items_parsed.append({
                    "supplier": row[col_map['supplier']].strip() if 'supplier' in col_map else '',
                    "date": safe_date(row[col_map['date']]) if 'date' in col_map else datetime.now(timezone.utc).strftime('%Y-%m-%d'),
                    "invoice_number": row[col_map['invoice_number']].strip() if 'invoice_number' in col_map else '',
                    "raw_name": item_name,
                    "quantity": qty,
                    "unit": row[col_map['unit']].strip() if 'unit' in col_map else '',
                    "unit_price": round(up, 2),
                    "total": round(tot, 2),
                })

            # Group into purchases by supplier+date+invoice
            groups = {}
            for it in items_parsed:
                key = (it['supplier'] or 'Unknown', it['date'], it['invoice_number'])
                groups.setdefault(key, []).append(it)

            if not groups and items_parsed:
                groups[('Unknown', items_parsed[0]['date'], '')] = items_parsed

            # If only one group or no grouping columns, return single purchase
            if len(groups) <= 1:
                all_items = [it for items in groups.values() for it in items]
                first = all_items[0] if all_items else {}
                subtotal = round(sum(it['total'] for it in all_items), 2)
                return {"extracted_data": {
                    "supplier_name": first.get('supplier', ''),
                    "invoice_date": first.get('date', datetime.now(timezone.utc).strftime('%Y-%m-%d')),
                    "invoice_number": first.get('invoice_number', ''),
                    "items": [{"raw_name": it['raw_name'], "quantity": it['quantity'], "unit": it['unit'], "unit_price": it['unit_price'], "total": it['total']} for it in all_items],
                    "subtotal": subtotal, "tax": 0, "total": subtotal,
                }, "document_type": document_type, "row_count": len(all_items)}
            else:
                # Multiple purchases - return first one and indicate more
                first_key = list(groups.keys())[0]
                first_items = groups[first_key]
                subtotal = round(sum(it['total'] for it in first_items), 2)
                return {"extracted_data": {
                    "supplier_name": first_key[0],
                    "invoice_date": first_key[1],
                    "invoice_number": first_key[2],
                    "items": [{"raw_name": it['raw_name'], "quantity": it['quantity'], "unit": it['unit'], "unit_price": it['unit_price'], "total": it['total']} for it in first_items],
                    "subtotal": subtotal, "tax": 0, "total": subtotal,
                }, "document_type": document_type, "row_count": len(items_parsed), "purchase_groups": len(groups),
                   "message": f"Found {len(groups)} purchases with {len(items_parsed)} total items. Showing the first purchase."}
        else:
            # Sales report
            items_parsed = []
            for row in data_rows:
                if len(row) <= max(col_map.values(), default=0):
                    row.extend([''] * (max(col_map.values(), default=0) + 1 - len(row)))
                item_name = row[col_map['item_name']].strip() if 'item_name' in col_map else ''
                if not item_name:
                    continue
                qty = safe_float(row[col_map['quantity']]) if 'quantity' in col_map else 1
                revenue = safe_float(row[col_map['total']]) if 'total' in col_map else 0
                items_parsed.append({"menu_item": item_name, "quantity": qty, "revenue": round(revenue, 2)})

            total_sales = round(sum(it['revenue'] for it in items_parsed), 2)
            report_date = datetime.now(timezone.utc).strftime('%Y-%m-%d')
            if 'date' in col_map and data_rows:
                report_date = safe_date(data_rows[0][col_map['date']])

            return {"extracted_data": {
                "report_date": report_date,
                "total_sales": total_sales,
                "items": items_parsed,
            }, "document_type": document_type, "row_count": len(items_parsed)}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Excel parse error: {e}")
        raise HTTPException(500, f"Failed to parse file: {str(e)}")

@api_router.post("/upload/extract")
async def extract_document(file: UploadFile = File(...), document_type: str = Form(...), user=Depends(get_user)):
    try:
        content = await file.read()
        mime = file.content_type or "image/jpeg"
        fname = (file.filename or "").lower()

        # Handle PDF: render all pages (up to 5) at high resolution and combine
        if "pdf" in mime.lower() or fname.endswith(".pdf"):
            import fitz
            pdf_doc = fitz.open(stream=content, filetype="pdf")
            images_b64 = []
            for page_num in range(min(len(pdf_doc), 5)):
                page = pdf_doc[page_num]
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                img_bytes = pix.tobytes("png")
                images_b64.append(base64.b64encode(img_bytes).decode())
            pdf_doc.close()
        else:
            images_b64 = [base64.b64encode(content).decode()]

        from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent

        if document_type == "purchase_invoice":
            prompt = """You are reading a restaurant purchase invoice or receipt. Extract ALL data into this exact JSON format:
{"supplier_name":"","invoice_date":"YYYY-MM-DD","invoice_number":"","items":[{"raw_name":"","quantity":0,"unit":"","unit_price":0,"total":0}],"subtotal":0,"tax":0,"total":0}

Rules:
- For each line item, calculate total = quantity × unit_price if not shown
- If unit_price is missing but total and quantity are known, calculate unit_price = total / quantity
- subtotal = sum of all item totals
- total = subtotal + tax
- Dates must be in YYYY-MM-DD format
- Use 0 for any truly missing numeric values
- Return ONLY the JSON object, no other text."""
        else:
            prompt = """You are reading a restaurant sales report or receipt. Extract ALL data into this exact JSON format:
{"report_date":"YYYY-MM-DD","total_sales":0,"items":[{"menu_item":"","quantity":0,"revenue":0}]}

Rules:
- total_sales should be the grand total
- For each item, revenue is the total amount for that item
- Dates must be in YYYY-MM-DD format
- Use 0 for any truly missing numeric values
- Return ONLY the JSON object, no other text."""

        chat = LlmChat(api_key=LLM_KEY, session_id=f"extract-{uuid.uuid4()}", system_message="You are an expert at reading restaurant invoices and receipts. Extract data accurately. Return valid JSON only, no markdown fences.").with_model("openai", "gpt-5.2")
        # Send all page images
        file_contents = [ImageContent(image_base64=b64) for b64 in images_b64]
        user_msg = UserMessage(text=prompt, file_contents=file_contents)
        response = await chat.send_message(user_msg)

        json_match = re.search(r'\{[\s\S]*\}', response)
        if json_match:
            extracted = json.loads(json_match.group())
        else:
            extracted = {"error": "Could not parse extraction results"}

        # Post-process: validate and fix calculations
        if "error" not in extracted:
            if document_type == "purchase_invoice":
                for item in extracted.get("items", []):
                    qty = float(item.get("quantity", 0) or 0)
                    up = float(item.get("unit_price", 0) or 0)
                    tot = float(item.get("total", 0) or 0)
                    if tot == 0 and qty > 0 and up > 0:
                        item["total"] = round(qty * up, 2)
                    elif up == 0 and tot > 0 and qty > 0:
                        item["unit_price"] = round(tot / qty, 2)
                    elif qty == 0 and tot > 0 and up > 0:
                        item["quantity"] = round(tot / up, 2)
                items_sum = round(sum(float(it.get("total", 0) or 0) for it in extracted.get("items", [])), 2)
                if not extracted.get("subtotal") and items_sum > 0:
                    extracted["subtotal"] = items_sum
                if not extracted.get("total") and items_sum > 0:
                    extracted["total"] = round(items_sum + float(extracted.get("tax", 0) or 0), 2)

        return {"extracted_data": extracted, "document_type": document_type}
    except Exception as e:
        logger.error(f"Extraction error: {e}")
        raise HTTPException(500, f"Extraction failed: {str(e)}")


# ==================== RECORDS LIBRARY ====================

UPLOADS_DIR = ROOT_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)

@api_router.post("/records/upload")
async def upload_record(
    file: UploadFile = File(...),
    folder: str = Form(...),
    transaction_type: str = Form(""),
    transaction_id: str = Form(""),
    transaction_date: str = Form(""),
    transaction_amount: float = Form(0),
    transaction_notes: str = Form(""),
    vendor_name: str = Form(""),
    user=Depends(get_user)
):
    """Upload a file to the Records Library."""
    import hashlib
    rid = user["restaurant_id"]
    if folder not in ("sales", "expenses"):
        raise HTTPException(400, "folder must be 'sales' or 'expenses'")

    content = await file.read()
    file_size = len(content)
    original_name = file.filename or "untitled"
    ext = original_name.rsplit(".", 1)[-1].lower() if "." in original_name else ""
    mime = file.content_type or "application/octet-stream"
    file_hash = hashlib.sha256(content).hexdigest()

    # Duplicate detection: same content hash OR same file name + size in same folder
    dup = await db.records_library.find_one({
        "restaurant_id": rid, "folder": folder,
        "$or": [
            {"file_hash": file_hash},
            {"file_name": original_name, "file_size": file_size},
        ]
    }, {"_id": 0, "id": 1, "file_name": 1, "upload_date": 1})
    if dup:
        raise HTTPException(
            409,
            f"Duplicate file detected: \"{dup['file_name']}\" (uploaded {dup.get('upload_date', 'previously')})"
        )

    record_id = str(uuid.uuid4())
    stored_name = f"{record_id}.{ext}" if ext else record_id

    # Save to disk
    file_path = UPLOADS_DIR / stored_name
    with open(file_path, "wb") as f:
        f.write(content)

    doc = {
        "id": record_id,
        "restaurant_id": rid,
        "folder": folder,
        "file_name": original_name,
        "file_type": mime,
        "file_extension": ext,
        "file_size": file_size,
        "file_hash": file_hash,
        "stored_name": stored_name,
        "upload_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "transaction_type": transaction_type,
        "transaction_id": transaction_id,
        "transaction_date": transaction_date,
        "transaction_amount": transaction_amount,
        "transaction_notes": transaction_notes,
        "vendor_name": vendor_name,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.records_library.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.get("/records")
async def list_records(
    user=Depends(get_user),
    folder: str = "",
    search: str = "",
    date_from: str = "",
    date_to: str = "",
    file_type: str = "",
    expense_category: str = "",
    sort_by: str = "upload_date",
    sort_order: str = "desc",
):
    """List records in the library with optional filters and sorting."""
    rid = user["restaurant_id"]
    query = {"restaurant_id": rid}
    if folder:
        query["folder"] = folder
    if search:
        query["file_name"] = {"$regex": search, "$options": "i"}
    if date_from:
        query.setdefault("upload_date", {})["$gte"] = date_from
    if date_to:
        query.setdefault("upload_date", {})["$lte"] = date_to
    if file_type and file_type != "all":
        if file_type == "image":
            query["file_type"] = {"$regex": "^image/", "$options": "i"}
        elif file_type == "pdf":
            query["file_extension"] = "pdf"
        elif file_type == "excel":
            query["file_extension"] = {"$in": ["xlsx", "xls", "csv"]}
    if expense_category and expense_category != "all":
        query["transaction_type"] = expense_category

    sort_field_map = {
        "upload_date": "upload_date",
        "amount": "transaction_amount",
        "name": "file_name",
    }
    sort_f = sort_field_map.get(sort_by, "upload_date")
    sort_d = -1 if sort_order == "desc" else 1

    records = await db.records_library.find(query, {"_id": 0}).sort(sort_f, sort_d).to_list(5000)
    return records


@api_router.get("/records/{record_id}")
async def get_record(record_id: str, user=Depends(get_user)):
    """Get a single record's details."""
    rec = await db.records_library.find_one(
        {"id": record_id, "restaurant_id": user["restaurant_id"]}, {"_id": 0}
    )
    if not rec:
        raise HTTPException(404, "Record not found")
    return rec


@api_router.get("/records/{record_id}/file")
async def serve_record_file(record_id: str, user=Depends(get_user)):
    """Serve the actual file for preview or download."""
    rec = await db.records_library.find_one(
        {"id": record_id, "restaurant_id": user["restaurant_id"]}, {"_id": 0}
    )
    if not rec:
        raise HTTPException(404, "Record not found")
    file_path = UPLOADS_DIR / rec["stored_name"]
    if not file_path.exists():
        raise HTTPException(404, "File not found on disk")
    content = file_path.read_bytes()
    return Response(
        content=content,
        media_type=rec.get("file_type", "application/octet-stream"),
        headers={"Content-Disposition": f'inline; filename="{rec["file_name"]}"'}
    )


@api_router.delete("/records/{record_id}")
async def delete_record(record_id: str, user=Depends(get_user)):
    """Delete a record and its file."""
    rec = await db.records_library.find_one(
        {"id": record_id, "restaurant_id": user["restaurant_id"]}, {"_id": 0}
    )
    if not rec:
        raise HTTPException(404, "Record not found")
    file_path = UPLOADS_DIR / rec["stored_name"]
    if file_path.exists():
        file_path.unlink()
    await db.records_library.delete_one({"id": record_id, "restaurant_id": user["restaurant_id"]})
    return {"status": "deleted"}


# ==================== DUPLICATE DETECTION ====================

class DuplicateCheckRequest(BaseModel):
    record_type: str  # "purchase", "sale", "salary", "other_expense"
    data: Dict[str, Any]

@api_router.post("/duplicates/check")
async def check_duplicates(req: DuplicateCheckRequest, user=Depends(get_user)):
    """Check for possible duplicate records before saving."""
    rid = user["restaurant_id"]
    rt = req.record_type
    d = req.data
    matches = []

    if rt == "purchase":
        query = {"restaurant_id": rid}
        # Check by invoice number (strongest signal)
        inv_no = d.get("invoice_number", "").strip()
        if inv_no:
            existing = await db.purchases.find(
                {**query, "invoice_number": {"$regex": f"^{inv_no}$", "$options": "i"}}, {"_id": 0, "id": 1, "supplier_name": 1, "invoice_number": 1, "invoice_date": 1, "total": 1}
            ).to_list(10)
            for e in existing:
                matches.append({"reason": f"Same invoice number: {inv_no}", "match_type": "invoice_number", **e})

        # Check by vendor + date + total
        vendor = d.get("supplier_name", "").strip()
        inv_date = d.get("invoice_date", "")
        total = d.get("total", 0)
        if vendor and inv_date and total:
            existing = await db.purchases.find(
                {**query, "supplier_name": {"$regex": f"^{vendor}$", "$options": "i"}, "invoice_date": inv_date, "total": {"$gte": total * 0.99, "$lte": total * 1.01}},
                {"_id": 0, "id": 1, "supplier_name": 1, "invoice_number": 1, "invoice_date": 1, "total": 1}
            ).to_list(10)
            for e in existing:
                if not any(m.get("id") == e.get("id") for m in matches):
                    matches.append({"reason": f"Same vendor ({vendor}), date ({inv_date}), and amount (${total:.2f})", "match_type": "vendor_date_amount", **e})

    elif rt == "sale":
        report_date = d.get("report_date", "")
        total_sales = d.get("total_sales", 0)
        if report_date:
            # Same date + same total
            existing = await db.sales.find(
                {"restaurant_id": rid, "report_date": report_date, "total_sales": {"$gte": total_sales * 0.99, "$lte": total_sales * 1.01}},
                {"_id": 0, "id": 1, "report_date": 1, "total_sales": 1}
            ).to_list(10)
            for e in existing:
                matches.append({"reason": f"Same date ({report_date}) and total (${total_sales:.2f})", "match_type": "date_amount", **e})

            # Same date only (weaker signal)
            if not matches:
                existing = await db.sales.find(
                    {"restaurant_id": rid, "report_date": report_date},
                    {"_id": 0, "id": 1, "report_date": 1, "total_sales": 1}
                ).to_list(10)
                for e in existing:
                    matches.append({"reason": f"A sales record already exists for {report_date}", "match_type": "date_only", **e})

    elif rt == "salary":
        employee = d.get("employee_name", "").strip()
        pay_date = d.get("payment_date", "")
        amount = d.get("amount", 0)
        if employee and pay_date:
            existing = await db.salaries.find(
                {"restaurant_id": rid, "employee_name": {"$regex": f"^{employee}$", "$options": "i"}, "payment_date": pay_date},
                {"_id": 0, "id": 1, "employee_name": 1, "payment_date": 1, "amount": 1, "position": 1}
            ).to_list(10)
            for e in existing:
                matches.append({"reason": f"Same employee ({employee}) and date ({pay_date})", "match_type": "employee_date", **e})

    elif rt == "other_expense":
        title = d.get("title", "").strip()
        exp_date = d.get("expense_date", "")
        amount = d.get("amount", 0)
        if title and exp_date:
            existing = await db.other_expenses.find(
                {"restaurant_id": rid, "title": {"$regex": f"^{title}$", "$options": "i"}, "expense_date": exp_date},
                {"_id": 0, "id": 1, "title": 1, "expense_date": 1, "amount": 1, "category": 1}
            ).to_list(10)
            for e in existing:
                matches.append({"reason": f"Same title ({title}) and date ({exp_date})", "match_type": "title_date", **e})

        if not matches and amount and exp_date:
            existing = await db.other_expenses.find(
                {"restaurant_id": rid, "expense_date": exp_date, "amount": {"$gte": amount * 0.99, "$lte": amount * 1.01}},
                {"_id": 0, "id": 1, "title": 1, "expense_date": 1, "amount": 1, "category": 1}
            ).to_list(10)
            for e in existing:
                if not any(m.get("id") == e.get("id") for m in matches):
                    matches.append({"reason": f"Same date ({exp_date}) and amount (${amount:.2f})", "match_type": "date_amount", **e})

    return {"has_duplicates": len(matches) > 0, "matches": matches}

# ==================== PURCHASES CRUD ====================

@api_router.get("/purchases")
async def list_purchases(user=Depends(get_user), search: str = "", supplier: str = "", date_from: str = "", date_to: str = "", sort_by: str = "invoice_date", sort_order: str = "desc"):
    query = {"restaurant_id": user["restaurant_id"]}
    if search:
        query["$or"] = [{"supplier_name": {"$regex": search, "$options": "i"}}, {"invoice_number": {"$regex": search, "$options": "i"}}]
    if supplier:
        query["supplier_name"] = {"$regex": supplier, "$options": "i"}
    if date_from:
        query.setdefault("invoice_date", {})["$gte"] = date_from
    if date_to:
        query.setdefault("invoice_date", {})["$lte"] = date_to
    return await db.purchases.find(query, {"_id": 0}).sort(sort_by, -1 if sort_order == "desc" else 1).to_list(1000)

@api_router.get("/purchases/{pid}")
async def get_purchase(pid: str, user=Depends(get_user)):
    p = await db.purchases.find_one({"id": pid, "restaurant_id": user["restaurant_id"]}, {"_id": 0})
    if not p:
        raise HTTPException(404, "Not found")
    return p

@api_router.post("/purchases")
async def create_purchase(data: PurchaseCreate, user=Depends(get_user)):
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["restaurant_id"] = user["restaurant_id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["created_by_id"] = user["id"]
    doc["created_by_name"] = user.get("name", "")
    doc["approval_status"] = _compute_approval_status(user, doc.get("total", 0))
    await db.purchases.insert_one(doc)
    doc.pop("_id", None)

    # --- Auto-create vendor if new ---
    rid = user["restaurant_id"]
    supplier_name = doc.get("supplier_name", "").strip()
    if supplier_name:
        existing_vendor = await db.suppliers.find_one({
            "restaurant_id": rid,
            "name": {"$regex": f"^{re.escape(supplier_name)}$", "$options": "i"}
        })
        if not existing_vendor:
            vendor_doc = {
                "id": str(uuid.uuid4()),
                "restaurant_id": rid,
                "name": supplier_name,
                "contact_name": "", "phone": "", "email": "", "address": "",
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.suppliers.insert_one(vendor_doc)
            logger.info(f"Auto-created vendor: {supplier_name}")

    # --- Auto-create items if new ---
    for item in doc.get("items", []):
        raw_name = item.get("raw_name", "").strip()
        if not raw_name:
            continue
        existing_item = await db.canonical_items.find_one({
            "restaurant_id": rid,
            "name": {"$regex": f"^{re.escape(raw_name)}$", "$options": "i"}
        })
        if not existing_item:
            item_doc = {
                "id": str(uuid.uuid4()),
                "restaurant_id": rid,
                "name": raw_name,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.canonical_items.insert_one(item_doc)
            logger.info(f"Auto-created item: {raw_name}")

    # --- Generate price alerts for items with price increases ---
    existing = await db.purchases.find(
        {"restaurant_id": rid, "id": {"$ne": doc["id"]}},
        {"_id": 0, "supplier_name": 1, "invoice_date": 1, "items": 1}
    ).to_list(10000)

    # Build alias mapping for fuzzy matching
    canon_items = await db.canonical_items.find({"restaurant_id": rid}, {"_id": 0}).to_list(1000)
    alias_list = await db.item_aliases.find({"restaurant_id": rid}, {"_id": 0}).to_list(5000)
    name_to_group = {}
    for c in canon_items:
        group_key = c["name"].lower()
        name_to_group[group_key] = group_key
    for a in alias_list:
        for c in canon_items:
            if c["id"] == a["canonical_item_id"]:
                name_to_group[a["alias_name"].lower()] = c["name"].lower()
                break

    for item in doc.get("items", []):
        raw = item.get("raw_name", "").strip()
        new_price = float(item.get("unit_price", 0))
        if not raw or new_price <= 0:
            continue

        # Determine the group of names to match against
        group_key = name_to_group.get(raw.lower(), raw.lower())
        match_names = {group_key}
        for k, v in name_to_group.items():
            if v == group_key:
                match_names.add(k)
        match_names.add(raw.lower())

        # Find the most recent previous price for this item across all purchases
        prev_record = None
        for p in sorted(existing, key=lambda x: x.get("invoice_date", ""), reverse=True):
            for it in p.get("items", []):
                if it.get("raw_name", "").lower() in match_names and float(it.get("unit_price", 0)) > 0:
                    prev_record = {"price": float(it["unit_price"]), "vendor": p.get("supplier_name", "Unknown"), "date": p.get("invoice_date", "")}
                    break
            if prev_record:
                break

        if prev_record and new_price > prev_record["price"]:
            pct = round(((new_price - prev_record["price"]) / prev_record["price"]) * 100, 1)
            alert_doc = {
                "id": str(uuid.uuid4()),
                "restaurant_id": rid,
                "type": "price_increase",
                "severity": "high" if pct > 15 else "medium",
                "item_name": raw,
                "previous_price": round(prev_record["price"], 2),
                "new_price": round(new_price, 2),
                "change_pct": pct,
                "vendor": doc.get("supplier_name", "Unknown"),
                "previous_vendor": prev_record["vendor"],
                "invoice_date": doc.get("invoice_date", ""),
                "message": f"Price increase detected for {raw}.\nPrevious price: ${prev_record['price']:.2f}\nNew price: ${new_price:.2f}",
                "is_read": False,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.alerts.insert_one(alert_doc)

    return doc

@api_router.put("/purchases/{pid}")
async def update_purchase(pid: str, data: PurchaseUpdate, user=Depends(get_user)):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(400, "No data")
    result = await db.purchases.update_one({"id": pid, "restaurant_id": user["restaurant_id"]}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(404, "Not found")
    return await db.purchases.find_one({"id": pid}, {"_id": 0})

@api_router.delete("/purchases/{pid}")
async def delete_purchase(pid: str, user=Depends(get_user)):
    result = await db.purchases.delete_one({"id": pid, "restaurant_id": user["restaurant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(404, "Not found")
    return {"status": "deleted"}

# ==================== SALARIES CRUD ====================

@api_router.get("/salaries")
async def list_salaries(user=Depends(get_user), date_from: str = "", date_to: str = "", sort_by: str = "payment_date", sort_order: str = "desc"):
    query = {"restaurant_id": user["restaurant_id"]}
    if date_from:
        query.setdefault("payment_date", {})["$gte"] = date_from
    if date_to:
        query.setdefault("payment_date", {})["$lte"] = date_to
    return await db.salaries.find(query, {"_id": 0}).sort(sort_by, -1 if sort_order == "desc" else 1).to_list(1000)

@api_router.post("/salaries")
async def create_salary(data: SalaryCreate, user=Depends(get_user)):
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["restaurant_id"] = user["restaurant_id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["created_by_id"] = user["id"]
    doc["created_by_name"] = user.get("name", "")
    doc["approval_status"] = _compute_approval_status(user, doc.get("amount", 0))
    await db.salaries.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.delete("/salaries/{sid}")
async def delete_salary(sid: str, user=Depends(get_user)):
    result = await db.salaries.delete_one({"id": sid, "restaurant_id": user["restaurant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(404, "Not found")
    return {"status": "deleted"}

# ==================== OTHER EXPENSES CRUD ====================

@api_router.get("/other-expenses")
async def list_other_expenses(user=Depends(get_user), category: str = "", date_from: str = "", date_to: str = "", sort_by: str = "expense_date", sort_order: str = "desc"):
    query = {"restaurant_id": user["restaurant_id"]}
    if category:
        query["category"] = category
    if date_from:
        query.setdefault("expense_date", {})["$gte"] = date_from
    if date_to:
        query.setdefault("expense_date", {})["$lte"] = date_to
    return await db.other_expenses.find(query, {"_id": 0}).sort(sort_by, -1 if sort_order == "desc" else 1).to_list(1000)

@api_router.post("/other-expenses")
async def create_other_expense(data: OtherExpenseCreate, user=Depends(get_user)):
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["restaurant_id"] = user["restaurant_id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["created_by_id"] = user["id"]
    doc["created_by_name"] = user.get("name", "")
    doc["approval_status"] = _compute_approval_status(user, doc.get("amount", 0))
    await db.other_expenses.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.delete("/other-expenses/{eid}")
async def delete_other_expense(eid: str, user=Depends(get_user)):
    result = await db.other_expenses.delete_one({"id": eid, "restaurant_id": user["restaurant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(404, "Not found")
    return {"status": "deleted"}

# ==================== SALES CRUD ====================

@api_router.get("/sales")
async def list_sales(user=Depends(get_user), search: str = "", date_from: str = "", date_to: str = "", sort_by: str = "report_date", sort_order: str = "desc"):
    query = {"restaurant_id": user["restaurant_id"]}
    if date_from:
        query.setdefault("report_date", {})["$gte"] = date_from
    if date_to:
        query.setdefault("report_date", {})["$lte"] = date_to
    return await db.sales.find(query, {"_id": 0}).sort(sort_by, -1 if sort_order == "desc" else 1).to_list(1000)

@api_router.get("/sales/{sid}")
async def get_sale(sid: str, user=Depends(get_user)):
    s = await db.sales.find_one({"id": sid, "restaurant_id": user["restaurant_id"]}, {"_id": 0})
    if not s:
        raise HTTPException(404, "Not found")
    return s

@api_router.post("/sales")
async def create_sale(data: SalesCreate, user=Depends(get_user)):
    doc = data.model_dump()
    # Normalize dates: date_from/date_to take precedence, fall back to report_date
    if doc.get("date_from") and doc.get("date_to"):
        if doc["date_to"] < doc["date_from"]:
            raise HTTPException(400, "To Date cannot be earlier than From Date")
        doc["report_date"] = doc["date_from"]
        doc["is_single_day"] = doc["date_from"] == doc["date_to"]
    elif doc.get("report_date"):
        doc["date_from"] = doc["report_date"]
        doc["date_to"] = doc["report_date"]
        doc["is_single_day"] = True
    doc["id"] = str(uuid.uuid4())
    doc["restaurant_id"] = user["restaurant_id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["created_by_id"] = user["id"]
    doc["created_by_name"] = user.get("name", "")
    doc["approval_status"] = _compute_approval_status(user, doc.get("total_sales", 0))
    await db.sales.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/sales/{sid}")
async def update_sale(sid: str, data: SalesUpdate, user=Depends(get_user)):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(400, "No data")
    await db.sales.update_one({"id": sid, "restaurant_id": user["restaurant_id"]}, {"$set": update_data})
    return await db.sales.find_one({"id": sid}, {"_id": 0})

@api_router.delete("/sales/{sid}")
async def delete_sale(sid: str, user=Depends(get_user)):
    result = await db.sales.delete_one({"id": sid, "restaurant_id": user["restaurant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(404, "Not found")
    return {"status": "deleted"}

# ==================== SUPPLIERS CRUD ====================

@api_router.get("/suppliers")
async def list_suppliers(user=Depends(get_user), search: str = ""):
    query = {"restaurant_id": user["restaurant_id"]}
    if search:
        query["name"] = {"$regex": search, "$options": "i"}
    suppliers = await db.suppliers.find(query, {"_id": 0}).to_list(1000)
    purchases = await db.purchases.find({"restaurant_id": user["restaurant_id"]}, {"_id": 0, "supplier_name": 1, "total": 1}).to_list(10000)
    for s in suppliers:
        s["total_spending"] = round(sum(p["total"] for p in purchases if p.get("supplier_name") == s["name"]), 2)
        s["invoice_count"] = sum(1 for p in purchases if p.get("supplier_name") == s["name"])
    return suppliers

@api_router.post("/suppliers")
async def create_supplier(data: SupplierCreate, user=Depends(get_user)):
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["restaurant_id"] = user["restaurant_id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.suppliers.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/suppliers/{sid}")
async def update_supplier(sid: str, data: SupplierCreate, user=Depends(get_user)):
    await db.suppliers.update_one({"id": sid, "restaurant_id": user["restaurant_id"]}, {"$set": data.model_dump()})
    return await db.suppliers.find_one({"id": sid}, {"_id": 0})

@api_router.delete("/suppliers/{sid}")
async def delete_supplier(sid: str, user=Depends(get_user)):
    result = await db.suppliers.delete_one({"id": sid, "restaurant_id": user["restaurant_id"]})
    if result.deleted_count == 0:
        raise HTTPException(404, "Not found")
    return {"status": "deleted"}

@api_router.get("/suppliers/{sid}/detail")
async def supplier_detail(sid: str, user=Depends(get_user)):
    rid = user["restaurant_id"]
    supplier = await db.suppliers.find_one({"id": sid, "restaurant_id": rid}, {"_id": 0})
    if not supplier:
        raise HTTPException(404, "Vendor not found")
    # Case-insensitive match on supplier name for purchases
    name = supplier["name"]
    purchases = await db.purchases.find(
        {"restaurant_id": rid, "supplier_name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}},
        {"_id": 0}
    ).to_list(10000)
    supplier["total_spending"] = round(sum(p.get("total", 0) for p in purchases), 2)
    supplier["invoice_count"] = len(purchases)
    return supplier

@api_router.get("/suppliers/{sid}/purchases")
async def supplier_purchases(sid: str, user=Depends(get_user), search: str = "", date_from: str = "", date_to: str = ""):
    rid = user["restaurant_id"]
    supplier = await db.suppliers.find_one({"id": sid, "restaurant_id": rid}, {"_id": 0})
    if not supplier:
        raise HTTPException(404, "Vendor not found")
    name = supplier["name"]
    query = {"restaurant_id": rid, "supplier_name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}}
    if search:
        query["invoice_number"] = {"$regex": search, "$options": "i"}
    if date_from:
        query.setdefault("invoice_date", {})["$gte"] = date_from
    if date_to:
        query.setdefault("invoice_date", {})["$lte"] = date_to
    purchases = await db.purchases.find(query, {"_id": 0}).sort("invoice_date", -1).to_list(10000)
    return purchases

# ==================== ITEMS & ALIASES ====================

@api_router.get("/items")
async def list_items(user=Depends(get_user), search: str = ""):
    query = {"restaurant_id": user["restaurant_id"]}
    if search:
        query["name"] = {"$regex": search, "$options": "i"}
    items = await db.canonical_items.find(query, {"_id": 0}).to_list(1000)
    for item in items:
        item["aliases"] = await db.item_aliases.find({"canonical_item_id": item["id"], "restaurant_id": user["restaurant_id"]}, {"_id": 0}).to_list(100)
    return items

@api_router.post("/items")
async def create_item(data: CanonicalItemCreate, user=Depends(get_user)):
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["restaurant_id"] = user["restaurant_id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.canonical_items.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/items/{iid}")
async def update_item(iid: str, data: CanonicalItemCreate, user=Depends(get_user)):
    await db.canonical_items.update_one({"id": iid, "restaurant_id": user["restaurant_id"]}, {"$set": data.model_dump()})
    return await db.canonical_items.find_one({"id": iid}, {"_id": 0})

@api_router.delete("/items/{iid}")
async def delete_item(iid: str, user=Depends(get_user)):
    await db.canonical_items.delete_one({"id": iid, "restaurant_id": user["restaurant_id"]})
    await db.item_aliases.delete_many({"canonical_item_id": iid})
    return {"status": "deleted"}

@api_router.post("/aliases")
async def create_alias(data: ItemAliasCreate, user=Depends(get_user)):
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["restaurant_id"] = user["restaurant_id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.item_aliases.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.delete("/aliases/{aid}")
async def delete_alias(aid: str, user=Depends(get_user)):
    await db.item_aliases.delete_one({"id": aid, "restaurant_id": user["restaurant_id"]})
    return {"status": "deleted"}

# ==================== ITEM PRICE HISTORY ====================

@api_router.get("/items/{item_id}/price-history")
async def item_price_history(item_id: str, user=Depends(get_user)):
    """Get price history for a canonical item by scanning all purchases matching its name + aliases."""
    rid = user["restaurant_id"]
    item = await db.canonical_items.find_one({"id": item_id, "restaurant_id": rid}, {"_id": 0})
    if not item:
        raise HTTPException(404, "Item not found")

    # Collect all names to match: canonical name + all aliases
    names = {item["name"].lower()}
    aliases = await db.item_aliases.find({"canonical_item_id": item_id, "restaurant_id": rid}, {"_id": 0}).to_list(200)
    for a in aliases:
        names.add(a["alias_name"].lower())

    # Scan purchases for matching line items
    purchases = await db.purchases.find({"restaurant_id": rid}, {"_id": 0, "supplier_name": 1, "invoice_date": 1, "items": 1}).to_list(10000)

    records = []
    for p in purchases:
        vendor = p.get("supplier_name", "Unknown")
        date = p.get("invoice_date", "")
        for it in p.get("items", []):
            raw = it.get("raw_name", "").lower()
            if raw in names:
                records.append({
                    "vendor": vendor,
                    "date": date,
                    "unit_price": round(float(it.get("unit_price", 0)), 2),
                    "quantity": float(it.get("quantity", 0)),
                    "unit": it.get("unit", ""),
                    "raw_name": it.get("raw_name", ""),
                })

    records.sort(key=lambda x: x["date"])

    # Build trend data: average price per date
    date_prices = {}
    for r in records:
        date_prices.setdefault(r["date"], []).append(r["unit_price"])
    trend = [{"date": d, "avg_price": round(sum(ps) / len(ps), 2)} for d, ps in sorted(date_prices.items())]

    # Summary stats
    all_prices = [r["unit_price"] for r in records if r["unit_price"] > 0]
    summary = {
        "total_records": len(records),
        "avg_price": round(sum(all_prices) / len(all_prices), 2) if all_prices else 0,
        "min_price": round(min(all_prices), 2) if all_prices else 0,
        "max_price": round(max(all_prices), 2) if all_prices else 0,
        "vendors": list(set(r["vendor"] for r in records)),
    }

    return {"item_name": item["name"], "records": records, "trend": trend, "summary": summary}

# ==================== REPORTS ====================

def _parse_report_dates(report_type, date, now):
    if report_type == "weekly":
        start = datetime.strptime(date, "%Y-%m-%d") if date else now - timedelta(days=now.weekday())
        start_str = start.strftime("%Y-%m-%d")
        end_str = (start + timedelta(days=6)).strftime("%Y-%m-%d")
        prev_start = (start - timedelta(days=7)).strftime("%Y-%m-%d")
        prev_end = (start - timedelta(days=1)).strftime("%Y-%m-%d")
    elif report_type == "monthly":
        start = datetime.strptime(date[:7] + "-01", "%Y-%m-%d") if date else now.replace(day=1)
        start_str = start.strftime("%Y-%m-%d")
        end = (start.replace(month=start.month + 1, day=1) if start.month < 12 else start.replace(year=start.year + 1, month=1, day=1)) - timedelta(days=1)
        end_str = end.strftime("%Y-%m-%d")
        prev_month = start - timedelta(days=1)
        prev_start = prev_month.replace(day=1).strftime("%Y-%m-%d")
        prev_end = prev_month.strftime("%Y-%m-%d")
    elif report_type == "quarterly":
        if date:
            # date format: "2026-Q1" or "2026-1"
            parts = date.replace("Q", "").replace("q", "").split("-")
            year = int(parts[0])
            quarter = int(parts[1]) if len(parts) > 1 else ((now.month - 1) // 3 + 1)
        else:
            year = now.year
            quarter = (now.month - 1) // 3 + 1
        q_month = (quarter - 1) * 3 + 1
        start_str = f"{year}-{q_month:02d}-01"
        end_month = q_month + 2
        end_date = datetime(year, end_month, 1) + timedelta(days=31)
        end_date = end_date.replace(day=1) - timedelta(days=1)
        end_str = end_date.strftime("%Y-%m-%d")
        # Previous quarter
        prev_q = quarter - 1 if quarter > 1 else 4
        prev_year = year if quarter > 1 else year - 1
        prev_q_month = (prev_q - 1) * 3 + 1
        prev_start = f"{prev_year}-{prev_q_month:02d}-01"
        prev_end_month = prev_q_month + 2
        prev_end_date = datetime(prev_year, prev_end_month, 1) + timedelta(days=31)
        prev_end_date = prev_end_date.replace(day=1) - timedelta(days=1)
        prev_end = prev_end_date.strftime("%Y-%m-%d")
    else:  # yearly
        year = int(date) if date else now.year
        start_str, end_str = f"{year}-01-01", f"{year}-12-31"
        prev_start, prev_end = f"{year-1}-01-01", f"{year-1}-12-31"
    return start_str, end_str, prev_start, prev_end

async def _build_report(rid, report_type, date):
    now = datetime.now(timezone.utc)
    start_str, end_str, prev_start, prev_end = _parse_report_dates(report_type, date, now)

    # Only include approved records in reports (backwards-compatible: no approval_status = approved)
    _appr = {"$or": [{"approval_status": {"$exists": False}}, {"approval_status": "approved"}]}

    purchases = await db.purchases.find({"restaurant_id": rid, **_appr, "invoice_date": {"$gte": start_str, "$lte": end_str}}, {"_id": 0}).to_list(10000)
    sales = await db.sales.find({"restaurant_id": rid, **_appr, "report_date": {"$gte": start_str, "$lte": end_str}}, {"_id": 0}).to_list(10000)
    salaries_cur = await db.salaries.find({"restaurant_id": rid, **_appr, "payment_date": {"$gte": start_str, "$lte": end_str}}, {"_id": 0}).to_list(10000)
    other_exp_cur = await db.other_expenses.find({"restaurant_id": rid, **_appr, "expense_date": {"$gte": start_str, "$lte": end_str}}, {"_id": 0}).to_list(10000)

    prev_purchases = await db.purchases.find({"restaurant_id": rid, **_appr, "invoice_date": {"$gte": prev_start, "$lte": prev_end}}, {"_id": 0}).to_list(10000)
    prev_sales = await db.sales.find({"restaurant_id": rid, **_appr, "report_date": {"$gte": prev_start, "$lte": prev_end}}, {"_id": 0}).to_list(10000)
    salaries_prev = await db.salaries.find({"restaurant_id": rid, **_appr, "payment_date": {"$gte": prev_start, "$lte": prev_end}}, {"_id": 0}).to_list(10000)
    other_exp_prev = await db.other_expenses.find({"restaurant_id": rid, **_appr, "expense_date": {"$gte": prev_start, "$lte": prev_end}}, {"_id": 0}).to_list(10000)

    total_p = round(sum(p["total"] for p in purchases), 2)
    total_s = round(sum(s["total_sales"] for s in sales), 2)
    total_sal = round(sum(s["amount"] for s in salaries_cur), 2)
    total_oe = round(sum(e["amount"] for e in other_exp_cur), 2)
    total_expenses = round(total_p + total_sal + total_oe, 2)
    net_profit = round(total_s - total_expenses, 2)

    prev_p = round(sum(p["total"] for p in prev_purchases), 2)
    prev_s = round(sum(s["total_sales"] for s in prev_sales), 2)
    prev_sal = round(sum(s["amount"] for s in salaries_prev), 2)
    prev_oe = round(sum(e["amount"] for e in other_exp_prev), 2)
    prev_total_expenses = round(prev_p + prev_sal + prev_oe, 2)
    prev_net_profit = round(prev_s - prev_total_expenses, 2)

    sup_spend = {}
    sup_invoice_count = {}
    for p in purchases:
        n = p.get("supplier_name", "Unknown")
        sup_spend[n] = sup_spend.get(n, 0) + p["total"]
        sup_invoice_count[n] = sup_invoice_count.get(n, 0) + 1

    item_spend = {}
    for p in purchases:
        for it in p.get("items", []):
            n = it.get("raw_name", "Unknown")
            item_spend[n] = item_spend.get(n, 0) + float(it.get("total", 0))

    # Price changes: compare avg unit_price in current vs previous period
    def item_prices(plist):
        prices = {}
        for p in plist:
            for it in p.get("items", []):
                n = it.get("raw_name", "Unknown")
                price = float(it.get("unit_price", 0))
                qty = float(it.get("quantity", 0))
                if price > 0:
                    prices.setdefault(n, []).append(price)
        return {n: round(sum(v)/len(v), 2) for n, v in prices.items()}

    cur_prices = item_prices(purchases)
    prev_prices = item_prices(prev_purchases)
    price_changes = []
    for name, cur_p in cur_prices.items():
        prev_p_val = prev_prices.get(name)
        if prev_p_val and prev_p_val > 0:
            pct = round(((cur_p - prev_p_val) / prev_p_val) * 100, 1)
            if abs(pct) > 0:
                price_changes.append({"item": name, "current_price": cur_p, "previous_price": prev_p_val, "change_pct": pct})
    price_changes.sort(key=lambda x: -abs(x["change_pct"]))

    daily = {}
    for p in purchases:
        d = p.get("invoice_date", "")
        daily.setdefault(d, {"date": d, "purchases": 0, "sales": 0})
        daily[d]["purchases"] += p["total"]
    for s in sales:
        d = s.get("report_date", "")
        daily.setdefault(d, {"date": d, "purchases": 0, "sales": 0})
        daily[d]["sales"] += s["total_sales"]

    alerts = await db.alerts.find({"restaurant_id": rid}, {"_id": 0}).to_list(100)

    return {
        "report_type": report_type, "date_range": {"start": start_str, "end": end_str},
        "prev_date_range": {"start": prev_start, "end": prev_end},
        "total_purchases": total_p, "total_sales": total_s, "profit": round(total_s - total_p, 2),
        "prev_purchases": prev_p, "prev_sales": prev_s, "prev_profit": round(prev_s - prev_p, 2),
        "margin_pct": round((total_s - total_p) / total_s * 100, 1) if total_s > 0 else 0,
        # Tax reporting fields
        "total_salaries": total_sal, "total_other_expenses": total_oe,
        "total_expenses": total_expenses, "net_profit": net_profit,
        "prev_salaries": prev_sal, "prev_other_expenses": prev_oe,
        "prev_total_expenses": prev_total_expenses, "prev_net_profit": prev_net_profit,
        "net_margin_pct": round(net_profit / total_s * 100, 1) if total_s > 0 else 0,
        # Existing fields
        "spending_by_supplier": [{"name": n, "total": round(t, 2), "invoices": sup_invoice_count.get(n, 0)} for n, t in sorted(sup_spend.items(), key=lambda x: -x[1])],
        "top_items": [{"name": n, "total": round(t, 2)} for n, t in sorted(item_spend.items(), key=lambda x: -x[1])[:10]],
        "price_changes": price_changes[:20],
        "daily_breakdown": sorted(daily.values(), key=lambda x: x["date"]),
        "alerts": alerts, "purchase_count": len(purchases), "sales_count": len(sales)
    }

@api_router.get("/reports")
async def get_reports(user=Depends(get_user), report_type: str = "weekly", date: str = ""):
    return await _build_report(user["restaurant_id"], report_type, date)

# ==================== DETAILED CATEGORY REPORTS ====================

@api_router.get("/reports/category/{category}")
async def get_category_report(category: str, user=Depends(get_user), date_from: str = "", date_to: str = "", vendor: str = ""):
    """Get detailed report for a specific category with from/to date filtering."""
    rid = user["restaurant_id"]
    now = datetime.now(timezone.utc)
    if not date_from:
        date_from = now.strftime("%Y-%m-01")
    if not date_to:
        date_to = now.strftime("%Y-%m-%d")
    # Only include approved records in reports
    _appr = {"$or": [{"approval_status": {"$exists": False}}, {"approval_status": "approved"}]}

    if category == "sales":
        sales = await db.sales.find(
            {"restaurant_id": rid, **_appr, "$or": [
                {"report_date": {"$gte": date_from, "$lte": date_to}},
                {"date_from": {"$gte": date_from, "$lte": date_to}},
            ]}, {"_id": 0}
        ).sort("report_date", -1).to_list(5000)
        total = round(sum(s.get("total_sales", 0) for s in sales), 2)
        avg_per_entry = round(total / len(sales), 2) if sales else 0
        return {"category": "sales", "date_from": date_from, "date_to": date_to,
                "total_sales": total, "record_count": len(sales), "avg_per_entry": avg_per_entry, "records": sales}

    elif category == "raw_materials":
        purchases = await db.purchases.find(
            {"restaurant_id": rid, **_appr, "invoice_date": {"$gte": date_from, "$lte": date_to}}, {"_id": 0}
        ).sort("invoice_date", -1).to_list(5000)
        total = round(sum(p.get("total", 0) for p in purchases), 2)
        # Flatten items for itemized view
        all_items = []
        for p in purchases:
            for it in p.get("items", []):
                all_items.append({
                    "vendor": p.get("supplier_name", ""), "date": p.get("invoice_date", ""),
                    "invoice": p.get("invoice_number", ""), "item": it.get("raw_name", ""),
                    "quantity": it.get("quantity", 0), "unit": it.get("unit", ""),
                    "unit_price": it.get("unit_price", 0), "line_total": it.get("total", 0)
                })
        return {"category": "raw_materials", "date_from": date_from, "date_to": date_to,
                "total": total, "invoice_count": len(purchases), "items": all_items, "records": purchases}

    elif category == "salaries":
        salaries = await db.salaries.find(
            {"restaurant_id": rid, **_appr, "payment_date": {"$gte": date_from, "$lte": date_to}}, {"_id": 0}
        ).sort("payment_date", -1).to_list(5000)
        total = round(sum(s.get("amount", 0) for s in salaries), 2)
        return {"category": "salaries", "date_from": date_from, "date_to": date_to,
                "total": total, "record_count": len(salaries), "records": salaries}

    elif category == "other_expenses":
        expenses = await db.other_expenses.find(
            {"restaurant_id": rid, **_appr, "expense_date": {"$gte": date_from, "$lte": date_to}}, {"_id": 0}
        ).sort("expense_date", -1).to_list(5000)
        total = round(sum(e.get("amount", 0) for e in expenses), 2)
        # Group by category
        by_cat = {}
        for e in expenses:
            c = e.get("category", "Other")
            by_cat[c] = by_cat.get(c, 0) + e.get("amount", 0)
        breakdown = [{"category": k, "total": round(v, 2)} for k, v in sorted(by_cat.items(), key=lambda x: -x[1])]
        return {"category": "other_expenses", "date_from": date_from, "date_to": date_to,
                "total": total, "record_count": len(expenses), "records": expenses, "breakdown": breakdown}

    elif category == "vendor":
        query = {"restaurant_id": rid, **_appr, "invoice_date": {"$gte": date_from, "$lte": date_to}}
        if vendor:
            query["supplier_name"] = {"$regex": f"^{vendor}$", "$options": "i"}
        purchases = await db.purchases.find(query, {"_id": 0}).sort("invoice_date", -1).to_list(5000)
        total = round(sum(p.get("total", 0) for p in purchases), 2)
        items = []
        for p in purchases:
            for it in p.get("items", []):
                items.append({
                    "vendor": p.get("supplier_name", ""), "date": p.get("invoice_date", ""),
                    "item": it.get("raw_name", ""), "quantity": it.get("quantity", 0),
                    "unit": it.get("unit", ""), "price": it.get("unit_price", 0),
                    "total": it.get("total", 0)
                })
        # Get vendor list for dropdown
        all_vendors = await db.suppliers.find({"restaurant_id": rid}, {"_id": 0, "name": 1}).to_list(200)
        vendor_names = sorted(set([v["name"] for v in all_vendors]))
        return {"category": "vendor", "date_from": date_from, "date_to": date_to, "vendor": vendor or "All",
                "total": total, "invoice_count": len(purchases), "items": items, "records": purchases, "vendors": vendor_names}

    elif category == "profit":
        sales = await db.sales.find(
            {"restaurant_id": rid, **_appr, "$or": [
                {"report_date": {"$gte": date_from, "$lte": date_to}},
                {"date_from": {"$gte": date_from, "$lte": date_to}},
            ]}, {"_id": 0}
        ).to_list(5000)
        purchases = await db.purchases.find(
            {"restaurant_id": rid, **_appr, "invoice_date": {"$gte": date_from, "$lte": date_to}}, {"_id": 0}
        ).to_list(5000)
        salaries = await db.salaries.find(
            {"restaurant_id": rid, **_appr, "payment_date": {"$gte": date_from, "$lte": date_to}}, {"_id": 0}
        ).to_list(5000)
        other_exp = await db.other_expenses.find(
            {"restaurant_id": rid, **_appr, "expense_date": {"$gte": date_from, "$lte": date_to}}, {"_id": 0}
        ).to_list(5000)
        total_sales = round(sum(s.get("total_sales", 0) for s in sales), 2)
        raw_mat = round(sum(p.get("total", 0) for p in purchases), 2)
        sal = round(sum(s.get("amount", 0) for s in salaries), 2)
        oe = round(sum(e.get("amount", 0) for e in other_exp), 2)
        total_exp = round(raw_mat + sal + oe, 2)
        net_profit = round(total_sales - total_exp, 2)
        margin = round(net_profit / total_sales * 100, 1) if total_sales > 0 else 0
        return {"category": "profit", "date_from": date_from, "date_to": date_to,
                "total_sales": total_sales, "raw_materials": raw_mat, "salaries": sal,
                "other_expenses": oe, "total_expenses": total_exp, "net_profit": net_profit, "net_margin_pct": margin}

    raise HTTPException(400, f"Unknown category: {category}")

@api_router.get("/reports/category/{category}/export")
async def export_category_report(category: str, fmt: str = "excel", user=Depends(get_user), date_from: str = "", date_to: str = "", vendor: str = ""):
    """Export a category report as PDF or Excel."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    # Fetch data
    report = await get_category_report(category, user, date_from, date_to, vendor)
    df = report.get("date_from", "")
    dt = report.get("date_to", "")
    title_map = {"sales": "Sales Report", "raw_materials": "Raw Material Expense Report", "salaries": "Salary Report",
                 "other_expenses": "Other Expense Report", "vendor": "Vendor Purchase Report", "profit": "Profit Report"}
    title = title_map.get(category, "Report")

    if fmt == "pdf":
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#0f172a'), spaceAfter=4*mm)
        sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#94a3b8'), spaceAfter=6*mm)
        elements = [Paragraph(title, title_style), Paragraph(f"{df} to {dt}", sub_style)]

        if category == "sales":
            data = [["Date", "Total Sales"]]
            for r in report.get("records", []):
                data.append([r.get("report_date", r.get("date_from", "")), f"${r.get('total_sales',0):,.2f}"])
            data.append(["TOTAL", f"${report['total_sales']:,.2f}"])
        elif category == "raw_materials":
            data = [["Vendor", "Item", "Date", "Qty", "Price", "Total"]]
            for it in report.get("items", []):
                data.append([it["vendor"], it["item"], it["date"], str(it["quantity"]), f"${it['unit_price']:,.2f}", f"${it['line_total']:,.2f}"])
            data.append(["", "", "", "", "TOTAL", f"${report['total']:,.2f}"])
        elif category == "salaries":
            data = [["Employee", "Position", "Amount", "Date"]]
            for r in report.get("records", []):
                data.append([r.get("employee_name",""), r.get("position",""), f"${r.get('amount',0):,.2f}", r.get("payment_date","")])
            data.append(["", "", f"${report['total']:,.2f}", "TOTAL"])
        elif category == "other_expenses":
            data = [["Title", "Category", "Amount", "Date", "Notes"]]
            for r in report.get("records", []):
                data.append([r.get("title",""), r.get("category",""), f"${r.get('amount',0):,.2f}", r.get("expense_date",""), (r.get("notes","") or "")[:30]])
            data.append(["", "", f"${report['total']:,.2f}", "TOTAL", ""])
        elif category == "vendor":
            data = [["Vendor", "Item", "Date", "Qty", "Price", "Total"]]
            for it in report.get("items", []):
                data.append([it["vendor"], it["item"], it["date"], str(it["quantity"]), f"${it['price']:,.2f}", f"${it['total']:,.2f}"])
            data.append(["", "", "", "", "TOTAL", f"${report['total']:,.2f}"])
        elif category == "profit":
            data = [["Category", "Amount"],
                    ["Total Sales", f"${report['total_sales']:,.2f}"], ["", ""],
                    ["Expenses:", ""], ["  Raw Materials", f"${report['raw_materials']:,.2f}"],
                    ["  Salaries", f"${report['salaries']:,.2f}"], ["  Other Expenses", f"${report['other_expenses']:,.2f}"],
                    ["Total Expenses", f"${report['total_expenses']:,.2f}"], ["", ""],
                    ["Net Profit", f"${report['net_profit']:,.2f}"], ["Net Margin", f"{report['net_margin_pct']}%"]]
        else:
            data = [["No data"]]

        t = RLTable(data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0f172a')), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTSIZE', (0,0), (-1,-1), 8), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
            ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'), ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#f1f5f9')),
        ]))
        elements.append(t)
        doc.build(elements)
        buf.seek(0)
        return Response(content=buf.read(), media_type="application/pdf",
                        headers={"Content-Disposition": f'attachment; filename="{category}_report_{df}_{dt}.pdf"'})

    else:  # excel
        wb = Workbook()
        ws = wb.active
        ws.title = title
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")

        ws.append([title])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append([f"Period: {df} to {dt}"])
        ws.append([])

        if category == "sales":
            headers = ["Date", "Total Sales"]
            ws.append(headers)
            for c in range(1, len(headers)+1): ws.cell(row=4, column=c).font = header_font; ws.cell(row=4, column=c).fill = header_fill
            for r in report.get("records", []):
                ws.append([r.get("report_date", r.get("date_from", "")), r.get("total_sales", 0)])
            ws.append(["TOTAL", report["total_sales"]])
        elif category == "raw_materials":
            headers = ["Vendor", "Item", "Date", "Quantity", "Unit Price", "Total"]
            ws.append(headers)
            for c in range(1, len(headers)+1): ws.cell(row=4, column=c).font = header_font; ws.cell(row=4, column=c).fill = header_fill
            for it in report.get("items", []):
                ws.append([it["vendor"], it["item"], it["date"], it["quantity"], it["unit_price"], it["line_total"]])
            ws.append(["", "", "", "", "TOTAL", report["total"]])
        elif category == "salaries":
            headers = ["Employee", "Position", "Amount", "Payment Date", "Notes"]
            ws.append(headers)
            for c in range(1, len(headers)+1): ws.cell(row=4, column=c).font = header_font; ws.cell(row=4, column=c).fill = header_fill
            for r in report.get("records", []):
                ws.append([r.get("employee_name",""), r.get("position",""), r.get("amount",0), r.get("payment_date",""), r.get("notes","")])
            ws.append(["", "", report["total"], "TOTAL", ""])
        elif category == "other_expenses":
            headers = ["Title", "Category", "Amount", "Date", "Notes"]
            ws.append(headers)
            for c in range(1, len(headers)+1): ws.cell(row=4, column=c).font = header_font; ws.cell(row=4, column=c).fill = header_fill
            for r in report.get("records", []):
                ws.append([r.get("title",""), r.get("category",""), r.get("amount",0), r.get("expense_date",""), r.get("notes","")])
            ws.append(["", "", report["total"], "TOTAL", ""])
        elif category == "vendor":
            headers = ["Vendor", "Item", "Date", "Quantity", "Price", "Total"]
            ws.append(headers)
            for c in range(1, len(headers)+1): ws.cell(row=4, column=c).font = header_font; ws.cell(row=4, column=c).fill = header_fill
            for it in report.get("items", []):
                ws.append([it["vendor"], it["item"], it["date"], it["quantity"], it["price"], it["total"]])
            ws.append(["", "", "", "", "TOTAL", report["total"]])
        elif category == "profit":
            headers = ["Category", "Amount"]
            ws.append(headers)
            for c in range(1, len(headers)+1): ws.cell(row=4, column=c).font = header_font; ws.cell(row=4, column=c).fill = header_fill
            ws.append(["Total Sales", report["total_sales"]])
            ws.append([])
            ws.append(["Raw Materials", report["raw_materials"]])
            ws.append(["Salaries", report["salaries"]])
            ws.append(["Other Expenses", report["other_expenses"]])
            ws.append(["Total Expenses", report["total_expenses"]])
            ws.append([])
            ws.append(["Net Profit", report["net_profit"]])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

        for col_cells in ws.columns:
            max_length = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_length + 4, 30)

        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(content=buf.read(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f'attachment; filename="{category}_report_{df}_{dt}.xlsx"'})

@api_router.get("/reports/download")
async def download_report(user=Depends(get_user), report_type: str = "weekly", date: str = "", fmt: str = "excel"):
    report = await _build_report(user["restaurant_id"], report_type, date)

    if fmt == "pdf":
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title2', parent=styles['Title'], fontSize=16, spaceAfter=6)
        sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=9, textColor=colors.grey)
        section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=12, spaceBefore=14, spaceAfter=6)

        elements = []
        elements.append(Paragraph("Restaurant Financial Report", title_style))
        elements.append(Paragraph(f"{report_type.title()} &bull; {report['date_range']['start']} to {report['date_range']['end']}", sub_style))
        elements.append(Spacer(1, 8*mm))

        # Tax Summary table
        elements.append(Paragraph("Tax Summary", section_style))
        tax_data = [
            ['Category', 'Amount'],
            ['Total Sales (Revenue)', f"${report['total_sales']:,.2f}"],
            ['', ''],
            ['Expenses Breakdown:', ''],
            ['  Raw Materials', f"${report['total_purchases']:,.2f}"],
            ['  Salaries', f"${report['total_salaries']:,.2f}"],
            ['  Other Expenses', f"${report['total_other_expenses']:,.2f}"],
            ['Total Expenses', f"${report['total_expenses']:,.2f}"],
            ['', ''],
            ['Net Profit', f"${report['net_profit']:,.2f}"],
            ['Net Margin', f"{report['net_margin_pct']}%"],
        ]
        t = Table(tax_data, colWidths=[100*mm, 60*mm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0f172a')), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTSIZE', (0,0), (-1,-1), 9), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
            ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('BACKGROUND', (0,3), (-1,3), colors.HexColor('#f8fafc')), ('FONTNAME', (0,3), (-1,3), 'Helvetica-Bold'),
            ('BACKGROUND', (0,7), (-1,7), colors.HexColor('#f1f5f9')), ('FONTNAME', (0,7), (-1,7), 'Helvetica-Bold'),
            ('BACKGROUND', (0,9), (-1,9), colors.HexColor('#ecfdf5') if report['net_profit'] >= 0 else colors.HexColor('#fef2f2')),
            ('FONTNAME', (0,9), (-1,9), 'Helvetica-Bold'), ('FONTNAME', (0,10), (-1,10), 'Helvetica-Bold'),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 4*mm))

        # KPIs table
        kpi_data = [['Revenue', 'Purchases', 'Profit', 'Margin'],
                     [f"${report['total_sales']:,.2f}", f"${report['total_purchases']:,.2f}", f"${report['profit']:,.2f}", f"{report['margin_pct']}%"]]
        t = Table(kpi_data, colWidths=[45*mm]*4)
        t.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0f172a')), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTSIZE', (0,0), (-1,-1), 9), ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
            ('TOPPADDING', (0,0), (-1,-1), 6), ('BOTTOMPADDING', (0,0), (-1,-1), 6)]))
        elements.append(t)

        # Supplier table
        if report['spending_by_supplier']:
            elements.append(Paragraph("Spending by Supplier", section_style))
            sup_data = [['Supplier', 'Total', 'Invoices']] + [[s['name'], f"${s['total']:,.2f}", str(s['invoices'])] for s in report['spending_by_supplier'][:10]]
            t = Table(sup_data, colWidths=[80*mm, 45*mm, 30*mm])
            t.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')), ('FONTSIZE', (0,0), (-1,-1), 8),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')), ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4)]))
            elements.append(t)

        # Price changes
        if report['price_changes']:
            elements.append(Paragraph("Price Changes", section_style))
            pc_data = [['Item', 'Previous', 'Current', 'Change']] + [[p['item'], f"${p['previous_price']:,.2f}", f"${p['current_price']:,.2f}", f"{p['change_pct']:+.1f}%"] for p in report['price_changes'][:15]]
            t = Table(pc_data, colWidths=[60*mm, 35*mm, 35*mm, 30*mm])
            t.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')), ('FONTSIZE', (0,0), (-1,-1), 8),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')), ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4)]))
            elements.append(t)

        doc.build(elements)
        buf.seek(0)
        filename = f"report_{report_type}_{report['date_range']['start']}.pdf"
        return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})

    else:  # excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        header_font = Font(bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        thin_border = Border(bottom=Side(style='thin', color='E2E8F0'))

        # Summary sheet
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Restaurant Financial Report"])
        ws['A1'].font = Font(bold=True, size=14)
        ws.append([f"{report_type.title()} Report: {report['date_range']['start']} to {report['date_range']['end']}"])
        ws.append([])
        ws.append(["Metric", "Current Period", "Previous Period", "Change"])
        for col in range(1, 5):
            cell = ws.cell(row=4, column=col)
            cell.font = header_font
            cell.fill = header_fill
        def pct_chg(cur, prev):
            return f"{((cur - prev) / prev * 100):+.1f}%" if prev else "N/A"
        ws.append(["Revenue", report['total_sales'], report['prev_sales'], pct_chg(report['total_sales'], report['prev_sales'])])
        ws.append(["Purchases", report['total_purchases'], report['prev_purchases'], pct_chg(report['total_purchases'], report['prev_purchases'])])
        ws.append(["Profit", report['profit'], report['prev_profit'], pct_chg(report['profit'], report['prev_profit']) if report['prev_profit'] != 0 else "N/A"])
        ws.append(["Margin %", f"{report['margin_pct']}%", "", ""])
        ws.append([])
        ws.append(["--- TAX SUMMARY ---"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
        ws.append(["Total Sales (Revenue)", report['total_sales']])
        ws.append([])
        ws.append(["Expenses Breakdown:"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(["  Raw Materials", report['total_purchases']])
        ws.append(["  Salaries", report['total_salaries']])
        ws.append(["  Other Expenses", report['total_other_expenses']])
        ws.append(["Total Expenses", report['total_expenses']])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append([])
        ws.append(["Net Profit", report['net_profit']])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
        ws.cell(row=ws.max_row, column=2).font = Font(bold=True, size=11)
        ws.append(["Net Margin", f"{report['net_margin_pct']}%"])
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 14

        # Suppliers sheet
        ws2 = wb.create_sheet("Suppliers")
        ws2.append(["Supplier", "Total Spent", "Invoices"])
        for col in range(1, 4):
            cell = ws2.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        for s in report['spending_by_supplier']:
            ws2.append([s['name'], s['total'], s['invoices']])
        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 12

        # Price changes sheet
        ws3 = wb.create_sheet("Price Changes")
        ws3.append(["Item", "Previous Price", "Current Price", "Change %"])
        for col in range(1, 5):
            cell = ws3.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        for p in report['price_changes']:
            ws3.append([p['item'], p['previous_price'], p['current_price'], p['change_pct']])
        ws3.column_dimensions['A'].width = 25
        ws3.column_dimensions['B'].width = 15
        ws3.column_dimensions['C'].width = 15
        ws3.column_dimensions['D'].width = 12

        # Daily breakdown
        ws4 = wb.create_sheet("Daily Breakdown")
        ws4.append(["Date", "Purchases", "Sales"])
        for col in range(1, 4):
            cell = ws4.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        for d in report['daily_breakdown']:
            ws4.append([d['date'], round(d['purchases'], 2), round(d['sales'], 2)])
        ws4.column_dimensions['A'].width = 15
        ws4.column_dimensions['B'].width = 15
        ws4.column_dimensions['C'].width = 15

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"report_{report_type}_{report['date_range']['start']}.xlsx"
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})

# ==================== PRICE INTELLIGENCE ====================

@api_router.get("/prices/intelligence")
async def price_intelligence(user=Depends(get_user)):
    """Comprehensive price tracking: supplier comparison, trends, and alerts."""
    rid = user["restaurant_id"]
    now = datetime.now(timezone.utc)
    purchases = await db.purchases.find({"restaurant_id": rid}, {"_id": 0}).to_list(10000)

    # --- 1. Per-item, per-supplier price tracking ---
    # Structure: {item_name: {supplier: [{price, date, qty}]}}
    item_supplier_prices = {}
    for p in purchases:
        supplier = p.get("supplier_name", "Unknown")
        inv_date = p.get("invoice_date", "")
        for it in p.get("items", []):
            name = it.get("raw_name", "Unknown")
            price = float(it.get("unit_price", 0))
            qty = float(it.get("quantity", 0))
            if price > 0:
                item_supplier_prices.setdefault(name, {}).setdefault(supplier, []).append({
                    "price": price, "date": inv_date, "qty": qty
                })

    # --- 2. Supplier comparison table ---
    # For each item: avg price per supplier, best supplier, potential savings
    all_suppliers = set()
    for p in purchases:
        sn = p.get("supplier_name", "").strip()
        if sn:
            all_suppliers.add(sn)
    all_suppliers = sorted(all_suppliers)

    comparison_items = []
    for item_name, suppliers_data in sorted(item_supplier_prices.items()):
        row = {"item": item_name, "suppliers": {}, "best_supplier": None, "best_price": None, "worst_price": None}
        for sup, entries in suppliers_data.items():
            avg_p = round(sum(e["price"] for e in entries) / len(entries), 2)
            latest_p = entries[-1]["price"] if entries else 0
            total_qty = sum(e["qty"] for e in entries)
            row["suppliers"][sup] = {"avg_price": avg_p, "latest_price": round(latest_p, 2), "purchase_count": len(entries), "total_qty": round(total_qty, 1)}
        # Find best/worst
        prices_list = [(sup, d["avg_price"]) for sup, d in row["suppliers"].items()]
        if prices_list:
            prices_list.sort(key=lambda x: x[1])
            row["best_supplier"] = prices_list[0][0]
            row["best_price"] = prices_list[0][1]
            row["worst_price"] = prices_list[-1][1]
            if len(prices_list) > 1 and prices_list[-1][1] > 0:
                # Savings if always bought from cheapest
                row["savings_pct"] = round((1 - prices_list[0][1] / prices_list[-1][1]) * 100, 1)
            else:
                row["savings_pct"] = 0
        comparison_items.append(row)

    # Sort by number of suppliers (multi-supplier items first), then by savings
    comparison_items.sort(key=lambda x: (-len(x["suppliers"]), -(x.get("savings_pct", 0))))

    # --- 3. Price trends over time ---
    # Group prices by item and date (weekly buckets)
    item_weekly = {}
    for item_name, suppliers_data in item_supplier_prices.items():
        all_entries = []
        for entries in suppliers_data.values():
            all_entries.extend(entries)
        all_entries.sort(key=lambda x: x["date"])

        weekly = {}
        for e in all_entries:
            if e["date"]:
                try:
                    d = datetime.strptime(e["date"], "%Y-%m-%d")
                    week_key = (d - timedelta(days=d.weekday())).strftime("%Y-%m-%d")
                    weekly.setdefault(week_key, []).append(e["price"])
                except ValueError:
                    pass
        if weekly:
            item_weekly[item_name] = [{"week": k, "avg_price": round(sum(v) / len(v), 2)} for k, v in sorted(weekly.items())]

    # Pick top 8 items by total spend for trend charts
    item_total_spend = {}
    for p in purchases:
        for it in p.get("items", []):
            name = it.get("raw_name", "Unknown")
            item_total_spend[name] = item_total_spend.get(name, 0) + float(it.get("total", 0))
    top_trend_items = sorted(item_total_spend.items(), key=lambda x: -x[1])[:8]
    price_trends = {name: item_weekly.get(name, []) for name, _ in top_trend_items if name in item_weekly}

    # --- 4. Price increase alerts (>10%) ---
    thirty_days_ago = (now - timedelta(days=30)).strftime("%Y-%m-%d")
    sixty_days_ago = (now - timedelta(days=60)).strftime("%Y-%m-%d")

    recent_avg = {}
    older_avg = {}
    for item_name, suppliers_data in item_supplier_prices.items():
        recent_prices = []
        older_prices = []
        for entries in suppliers_data.values():
            for e in entries:
                if e["date"] >= thirty_days_ago:
                    recent_prices.append(e["price"])
                elif e["date"] >= sixty_days_ago:
                    older_prices.append(e["price"])
        if recent_prices:
            recent_avg[item_name] = sum(recent_prices) / len(recent_prices)
        if older_prices:
            older_avg[item_name] = sum(older_prices) / len(older_prices)

    price_alerts = []
    for name, cur in recent_avg.items():
        prev = older_avg.get(name)
        if prev and prev > 0:
            pct = ((cur - prev) / prev) * 100
            if pct > 10:
                price_alerts.append({
                    "item": name,
                    "current_avg": round(cur, 2),
                    "previous_avg": round(prev, 2),
                    "change_pct": round(pct, 1),
                    "severity": "high" if pct > 20 else "medium",
                })
    price_alerts.sort(key=lambda x: -x["change_pct"])

    return {
        "suppliers": all_suppliers,
        "comparison": comparison_items[:30],
        "price_trends": price_trends,
        "price_alerts": price_alerts[:15],
        "total_items_tracked": len(item_supplier_prices),
        "total_suppliers": len(all_suppliers),
    }

# ==================== VENDOR PRICE COMPARISON ====================

@api_router.get("/prices/vendor-comparison")
async def vendor_price_comparison(user=Depends(get_user)):
    """Per-item, per-vendor latest price comparison with best vendor highlighted."""
    rid = user["restaurant_id"]
    purchases = await db.purchases.find({"restaurant_id": rid}, {"_id": 0, "supplier_name": 1, "invoice_date": 1, "items": 1}).to_list(10000)

    # Resolve canonical names: build alias->canonical mapping
    canonical = await db.canonical_items.find({"restaurant_id": rid}, {"_id": 0}).to_list(1000)
    aliases = await db.item_aliases.find({"restaurant_id": rid}, {"_id": 0}).to_list(5000)
    alias_to_canonical = {}
    for a in aliases:
        for c in canonical:
            if c["id"] == a["canonical_item_id"]:
                alias_to_canonical[a["alias_name"].lower()] = c["name"]
                break
    for c in canonical:
        alias_to_canonical[c["name"].lower()] = c["name"]

    # Structure: {canonical_name: {vendor: [{price, date, raw_name}]}}
    item_vendor_prices = {}
    for p in purchases:
        vendor = p.get("supplier_name", "Unknown")
        inv_date = p.get("invoice_date", "")
        for it in p.get("items", []):
            raw = it.get("raw_name", "")
            price = float(it.get("unit_price", 0))
            if price <= 0:
                continue
            canon = alias_to_canonical.get(raw.lower(), raw)
            item_vendor_prices.setdefault(canon, {}).setdefault(vendor, []).append({
                "price": price, "date": inv_date, "raw_name": raw,
                "quantity": float(it.get("quantity", 0)), "unit": it.get("unit", ""),
            })

    items_out = []
    for item_name, vendors_data in sorted(item_vendor_prices.items()):
        vendors_list = []
        for vendor_name, entries in sorted(vendors_data.items()):
            entries.sort(key=lambda x: x["date"], reverse=True)
            latest = entries[0]
            avg = round(sum(e["price"] for e in entries) / len(entries), 2)
            vendors_list.append({
                "vendor": vendor_name,
                "latest_price": round(latest["price"], 2),
                "latest_date": latest["date"],
                "avg_price": avg,
                "purchase_count": len(entries),
                "unit": latest.get("unit", ""),
            })

        vendors_list.sort(key=lambda x: x["latest_price"])
        best_vendor = vendors_list[0]["vendor"] if vendors_list else None
        best_price = vendors_list[0]["latest_price"] if vendors_list else 0
        worst_price = vendors_list[-1]["latest_price"] if len(vendors_list) > 1 else best_price
        savings_pct = round((1 - best_price / worst_price) * 100, 1) if worst_price > 0 and len(vendors_list) > 1 else 0

        items_out.append({
            "item": item_name,
            "vendors": vendors_list,
            "best_vendor": best_vendor,
            "best_price": best_price,
            "savings_pct": savings_pct,
            "vendor_count": len(vendors_list),
        })

    items_out.sort(key=lambda x: (-x["vendor_count"], -x["savings_pct"]))
    return {"items": items_out, "total_items": len(items_out)}

# ==================== ALERTS ====================

@api_router.get("/alerts/prices")
async def list_price_alerts(user=Depends(get_user)):
    """Get price increase alerts, newest first."""
    return await db.alerts.find(
        {"restaurant_id": user["restaurant_id"], "type": "price_increase"},
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)

@api_router.delete("/alerts/prices/{aid}")
async def dismiss_price_alert(aid: str, user=Depends(get_user)):
    result = await db.alerts.delete_one({"id": aid, "restaurant_id": user["restaurant_id"], "type": "price_increase"})
    if result.deleted_count == 0:
        raise HTTPException(404, "Not found")
    return {"status": "deleted"}

@api_router.get("/alerts")
async def list_alerts(user=Depends(get_user)):
    return await db.alerts.find({"restaurant_id": user["restaurant_id"]}, {"_id": 0}).sort("created_at", -1).to_list(100)

@api_router.put("/alerts/{aid}/read")
async def mark_alert_read(aid: str, user=Depends(get_user)):
    await db.alerts.update_one({"id": aid, "restaurant_id": user["restaurant_id"]}, {"$set": {"is_read": True}})
    return {"status": "ok"}

# ==================== SMART PURCHASE DECISIONS ====================

@api_router.get("/purchase-decisions")
async def smart_purchase_decisions(user=Depends(get_user)):
    """Generate smart purchase insights from real purchase data."""
    rid = user["restaurant_id"]
    now = datetime.now(timezone.utc)
    purchases = await db.purchases.find(
        {"restaurant_id": rid, "$or": [{"approval_status": {"$exists": False}}, {"approval_status": "approved"}]},
        {"_id": 0, "supplier_name": 1, "invoice_date": 1, "items": 1}
    ).to_list(10000)

    if not purchases:
        return {"items": [], "insights": [], "weekly_changes": [], "potential_savings": 0, "total_items": 0}

    this_week_start = (now - timedelta(days=now.weekday())).strftime("%Y-%m-%d")
    last_week_start = (now - timedelta(days=now.weekday() + 7)).strftime("%Y-%m-%d")
    last_week_end = (now - timedelta(days=now.weekday() + 1)).strftime("%Y-%m-%d")

    # Build per-item, per-vendor price history
    item_vendor_data = {}  # item -> vendor -> [{price, date, qty, unit}]
    for p in purchases:
        vendor = p.get("supplier_name", "")
        inv_date = p.get("invoice_date", "")
        for it in p.get("items", []):
            raw = it.get("raw_name", "").strip()
            price = float(it.get("unit_price", 0) or 0)
            qty = float(it.get("quantity", 0) or 0)
            if raw and price > 0:
                item_vendor_data.setdefault(raw, {}).setdefault(vendor, []).append({
                    "price": price, "date": inv_date, "qty": qty, "unit": it.get("unit", "")
                })

    items_out = []
    insights = []
    weekly_changes = []
    total_potential_savings = 0

    for item_name, vendors_data in sorted(item_vendor_data.items()):
        # Per vendor: latest price, average, count
        vendor_summaries = []
        all_prices = []
        for vendor_name, entries in vendors_data.items():
            entries.sort(key=lambda x: x["date"])
            latest = entries[-1]
            avg_price = round(sum(e["price"] for e in entries) / len(entries), 2)
            vendor_summaries.append({
                "vendor": vendor_name,
                "latest_price": round(latest["price"], 2),
                "avg_price": avg_price,
                "latest_date": latest["date"],
                "purchase_count": len(entries),
                "unit": latest.get("unit", ""),
            })
            all_prices.extend(entries)

        vendor_summaries.sort(key=lambda x: x["latest_price"])
        best = vendor_summaries[0]

        # Weekly comparison: this week vs last week prices
        this_week_prices = [e["price"] for e in all_prices if e["date"] >= this_week_start]
        last_week_prices = [e["price"] for e in all_prices if last_week_start <= e["date"] <= last_week_end]

        week_change = None
        if this_week_prices and last_week_prices:
            tw_avg = sum(this_week_prices) / len(this_week_prices)
            lw_avg = sum(last_week_prices) / len(last_week_prices)
            if lw_avg > 0:
                change_pct = round(((tw_avg - lw_avg) / lw_avg) * 100, 1)
                if abs(change_pct) > 1:
                    week_change = {
                        "item": item_name,
                        "this_week_avg": round(tw_avg, 2),
                        "last_week_avg": round(lw_avg, 2),
                        "change_pct": change_pct,
                        "direction": "up" if change_pct > 0 else "down",
                        "unit": best.get("unit", ""),
                    }
                    weekly_changes.append(week_change)

        # Savings: difference between best and worst vendor (per-unit)
        saving_per_unit = 0
        if len(vendor_summaries) > 1:
            worst = vendor_summaries[-1]
            saving_per_unit = round(worst["latest_price"] - best["latest_price"], 2)
            # Estimate total savings based on avg weekly quantity from worst vendor
            worst_entries = vendors_data.get(worst["vendor"], [])
            avg_qty = sum(e["qty"] for e in worst_entries) / max(len(worst_entries), 1)
            total_potential_savings += round(saving_per_unit * avg_qty, 2)

            if saving_per_unit > 0:
                insights.append({
                    "type": "best_vendor",
                    "item": item_name,
                    "best_vendor": best["vendor"],
                    "best_price": best["latest_price"],
                    "worst_vendor": worst["vendor"],
                    "worst_price": worst["latest_price"],
                    "saving_per_unit": saving_per_unit,
                    "unit": best.get("unit", ""),
                    "pct": round((saving_per_unit / worst["latest_price"]) * 100, 1) if worst["latest_price"] > 0 else 0,
                })

        items_out.append({
            "item": item_name,
            "vendors": vendor_summaries,
            "best_vendor": best["vendor"],
            "best_price": best["latest_price"],
            "saving_per_unit": saving_per_unit,
            "vendor_count": len(vendor_summaries),
            "unit": best.get("unit", ""),
            "week_change": week_change,
        })

    # Add price increase insights
    for wc in weekly_changes:
        if wc["direction"] == "up" and wc["change_pct"] > 3:
            insights.append({
                "type": "price_increase",
                "item": wc["item"],
                "change_pct": wc["change_pct"],
                "this_week": wc["this_week_avg"],
                "last_week": wc["last_week_avg"],
                "unit": wc.get("unit", ""),
            })

    # Sort insights by impact
    insights.sort(key=lambda x: -(x.get("saving_per_unit", 0) or x.get("change_pct", 0)))

    return {
        "items": items_out,
        "insights": insights[:20],
        "weekly_changes": sorted(weekly_changes, key=lambda x: -abs(x["change_pct"])),
        "potential_savings": round(total_potential_savings, 2),
        "total_items": len(items_out),
    }

# ==================== CHAT ====================

@api_router.get("/chat/messages")
async def get_chat_messages(user=Depends(get_user)):
    return await db.chat_messages.find({"user_id": user["id"]}, {"_id": 0}).sort("created_at", 1).to_list(100)

@api_router.post("/chat")
async def send_chat(data: ChatMessageIn, user=Depends(get_user)):
    rid = user["restaurant_id"]
    user_msg = {"id": str(uuid.uuid4()), "user_id": user["id"], "restaurant_id": rid, "role": "user", "content": data.message, "created_at": datetime.now(timezone.utc).isoformat()}
    await db.chat_messages.insert_one(user_msg)
    user_msg.pop("_id", None)

    purchases = await db.purchases.find({"restaurant_id": rid}, {"_id": 0}).to_list(10000)
    sales = await db.sales.find({"restaurant_id": rid}, {"_id": 0}).to_list(10000)
    suppliers = await db.suppliers.find({"restaurant_id": rid}, {"_id": 0}).to_list(100)
    alerts = await db.alerts.find({"restaurant_id": rid}, {"_id": 0}).to_list(50)
    now = datetime.now(timezone.utc)

    # Build period-based summaries
    week_start = (now - timedelta(days=now.weekday())).strftime("%Y-%m-%d")
    month_start = now.strftime("%Y-%m-01")
    year_start = now.strftime("%Y-01-01")
    prev_week_start = (now - timedelta(days=now.weekday() + 7)).strftime("%Y-%m-%d")
    prev_week_end = (now - timedelta(days=now.weekday() + 1)).strftime("%Y-%m-%d")
    prev_month = now.replace(day=1) - timedelta(days=1)
    prev_month_start = prev_month.strftime("%Y-%m-01")
    prev_month_end = prev_month.strftime("%Y-%m-%d")

    def sum_p(df, dt=None):
        return sum(p["total"] for p in purchases if p.get("invoice_date", "") >= df and (not dt or p.get("invoice_date", "") <= dt))
    def sum_s(df, dt=None):
        return sum(s["total_sales"] for s in sales if s.get("report_date", "") >= df and (not dt or s.get("report_date", "") <= dt))

    # Item price tracking for alerts
    item_prices = {}
    for p in sorted(purchases, key=lambda x: x.get("invoice_date", "")):
        for it in p.get("items", []):
            name = it.get("raw_name", "Unknown")
            price = float(it.get("unit_price", 0))
            if price > 0:
                item_prices.setdefault(name, []).append({"price": price, "date": p.get("invoice_date", "")})

    price_changes = []
    for name, prices in item_prices.items():
        if len(prices) >= 2:
            old, new = prices[-2]["price"], prices[-1]["price"]
            if old > 0:
                pct = ((new - old) / old) * 100
                if abs(pct) > 5:
                    price_changes.append({"item": name, "old": old, "new": new, "change_pct": round(pct, 1)})

    sup_spend = {}
    for p in purchases:
        n = p.get("supplier_name", "Unknown")
        sup_spend[n] = sup_spend.get(n, 0) + p["total"]
    top_suppliers = sorted(sup_spend.items(), key=lambda x: -x[1])[:5]

    item_spend = {}
    for p in purchases:
        for it in p.get("items", []):
            n = it.get("raw_name", "Unknown")
            item_spend[n] = item_spend.get(n, 0) + float(it.get("total", 0))
    top_items = sorted(item_spend.items(), key=lambda x: -x[1])[:10]

    # Smart alerts for chat context
    smart_alerts = await _generate_smart_alerts(rid)
    smart_alerts_ctx = ""
    if smart_alerts:
        sa_lines = []
        for sa in smart_alerts:
            sa_lines.append(f"- [{sa['type'].upper()}] {sa['title']} — {sa['detail']}")
        smart_alerts_ctx = "\nSMART ALERTS (active issues detected):\n" + "\n".join(sa_lines)

    context = f"""RESTAURANT FINANCIAL DATA (as of {now.strftime('%A, %B %d, %Y')}):

WEEKLY SNAPSHOT (This Week starting {week_start}):
- Purchases: ${sum_p(week_start):,.2f} | Sales: ${sum_s(week_start):,.2f}
- Gross Margin: ${sum_s(week_start) - sum_p(week_start):,.2f}

PREVIOUS WEEK ({prev_week_start} to {prev_week_end}):
- Purchases: ${sum_p(prev_week_start, prev_week_end):,.2f} | Sales: ${sum_s(prev_week_start, prev_week_end):,.2f}

MONTHLY SNAPSHOT (This Month starting {month_start}):
- Purchases: ${sum_p(month_start):,.2f} | Sales: ${sum_s(month_start):,.2f}
- Gross Margin: ${sum_s(month_start) - sum_p(month_start):,.2f}

PREVIOUS MONTH ({prev_month_start} to {prev_month_end}):
- Purchases: ${sum_p(prev_month_start, prev_month_end):,.2f} | Sales: ${sum_s(prev_month_start, prev_month_end):,.2f}

YEAR-TO-DATE ({year_start}):
- Total Purchases: ${sum_p(year_start):,.2f} | Total Sales: ${sum_s(year_start):,.2f}
- Net Margin: ${sum_s(year_start) - sum_p(year_start):,.2f}

ALL-TIME OVERVIEW:
- {len(purchases)} purchase invoices totaling ${sum(p['total'] for p in purchases):,.2f}
- {len(sales)} sales reports totaling ${sum(s['total_sales'] for s in sales):,.2f}
- {len(suppliers)} active suppliers

TOP SUPPLIERS BY SPEND: {json.dumps([{"name": n, "total": round(t, 2)} for n, t in top_suppliers])}
TOP ITEMS BY COST: {json.dumps([{"name": n, "total": round(t, 2)} for n, t in top_items])}
RECENT PRICE CHANGES (>5%): {json.dumps(price_changes[:10]) if price_changes else "None detected"}
ACTIVE ALERTS: {len([a for a in alerts if not a.get('is_read')])}
RECENT PURCHASES: {json.dumps([{'date': p['invoice_date'], 'supplier': p['supplier_name'], 'total': p['total']} for p in sorted(purchases, key=lambda x: x.get('invoice_date',''), reverse=True)[:8]])}
RECENT SALES: {json.dumps([{'date': s['report_date'], 'total': s['total_sales']} for s in sorted(sales, key=lambda x: x.get('report_date',''), reverse=True)[:8]])}
{smart_alerts_ctx}"""

    system_msg = f"""You are Restaurant Accountant AI, a senior financial analyst for restaurants. Follow these rules strictly:

1. STYLE: Be concise and data-driven. Lead with key numbers. Use short paragraphs and bullet points.
2. FORMAT: Use **bold** for key figures and metrics. Use bullet points for lists. Keep responses under 200 words unless the question requires detail.
3. PERIODS: Always reference specific time periods (this week, this month, YTD, etc.) when discussing finances. Compare to previous periods when relevant.
4. INSIGHTS: Don't just report numbers - provide brief actionable insights (e.g., "Spending is up 12% WoW - driven mainly by produce costs").
5. CURRENCY: Always format dollar amounts with commas (e.g., $1,234.56).
6. SMART ALERTS: You have access to a smart alerts system that detects low stock ingredients, cost increases, and profit margin drops. When users ask about stock, costs, margins, or alerts, reference the specific alert data provided. Be proactive about mentioning relevant alerts even if the user doesn't ask directly.
7. If you don't have enough data to answer precisely, say so honestly and suggest what data would help.

{context}"""

    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage as LlmUserMessage
        chat = LlmChat(api_key=LLM_KEY, session_id=f"chat-{user['id']}-{uuid.uuid4()}", system_message=system_msg).with_model("openai", "gpt-5.2")
        response = await chat.send_message(LlmUserMessage(text=data.message))
    except Exception as e:
        logger.error(f"Chat error: {e}")
        response = f"I'm having trouble connecting to the AI service right now. Please try again later."

    asst_msg = {"id": str(uuid.uuid4()), "user_id": user["id"], "restaurant_id": rid, "role": "assistant", "content": response, "created_at": datetime.now(timezone.utc).isoformat()}
    await db.chat_messages.insert_one(asst_msg)
    asst_msg.pop("_id", None)
    return {"user_message": user_msg, "assistant_message": asst_msg}

@api_router.delete("/chat/messages")
async def clear_chat(user=Depends(get_user)):
    await db.chat_messages.delete_many({"user_id": user["id"]})
    return {"status": "cleared"}

# ==================== SETTINGS ====================

SETTINGS_DEFAULTS = {
    "currency": "USD", "default_tax_rate": 0, "default_expense_category": "Rent",
    "alerts_enabled": True, "alert_price_increase": True, "alert_cheaper_vendor": True, "alert_not_ordered": True,
    "language": "en", "date_format": "YYYY-MM-DD",
}

@api_router.get("/settings")
async def get_settings(user=Depends(get_user)):
    r = await db.restaurants.find_one({"id": user["restaurant_id"]}, {"_id": 0}) or {}
    # Merge defaults for any missing settings fields
    settings = {k: r.get(k, v) for k, v in SETTINGS_DEFAULTS.items()}
    return {
        "user": {"id": user["id"], "email": user["email"], "name": user["name"]},
        "restaurant": {**r, **settings},
    }

@api_router.put("/settings")
async def update_settings(data: SettingsUpdate, user=Depends(get_user)):
    if data.name:
        await db.users.update_one({"id": user["id"]}, {"$set": {"name": data.name}})
    rid = user["restaurant_id"]
    update_fields = {}
    if data.restaurant_name is not None: update_fields["name"] = data.restaurant_name
    if data.address is not None: update_fields["address"] = data.address
    if data.phone is not None: update_fields["phone"] = data.phone
    if data.email is not None: update_fields["email"] = data.email
    if data.currency is not None: update_fields["currency"] = data.currency
    if data.default_tax_rate is not None: update_fields["default_tax_rate"] = data.default_tax_rate
    if data.default_expense_category is not None: update_fields["default_expense_category"] = data.default_expense_category
    if data.alerts_enabled is not None: update_fields["alerts_enabled"] = data.alerts_enabled
    if data.alert_price_increase is not None: update_fields["alert_price_increase"] = data.alert_price_increase
    if data.alert_cheaper_vendor is not None: update_fields["alert_cheaper_vendor"] = data.alert_cheaper_vendor
    if data.alert_not_ordered is not None: update_fields["alert_not_ordered"] = data.alert_not_ordered
    if data.language is not None: update_fields["language"] = data.language
    if data.date_format is not None: update_fields["date_format"] = data.date_format
    if update_fields:
        await db.restaurants.update_one({"id": rid}, {"$set": update_fields})
    # Re-fetch user to get updated name
    updated_user = await db.users.find_one({"id": user["id"]}, {"_id": 0})
    return await get_settings(updated_user)

@api_router.post("/settings/reset-data")
async def reset_all_data(user=Depends(get_user)):
    rid = user["restaurant_id"]
    for coll_name in ["purchases", "sales", "salaries", "other_expenses", "suppliers", "canonical_items", "item_aliases", "alerts", "records_library"]:
        await db[coll_name].delete_many({"restaurant_id": rid})
    # Also clear data without restaurant_id scoping (chat messages)
    await db.chat_messages.delete_many({"user_id": user["id"]})
    return {"status": "All data has been reset"}

@api_router.post("/settings/upload-logo")
async def upload_logo(file: UploadFile = File(...), user=Depends(get_user)):
    content = await file.read()
    if len(content) > 2 * 1024 * 1024:
        raise HTTPException(400, "Logo must be under 2MB")
    b64 = base64.b64encode(content).decode()
    mime = file.content_type or "image/png"
    data_url = f"data:{mime};base64,{b64}"
    await db.restaurants.update_one({"id": user["restaurant_id"]}, {"$set": {"logo": data_url}})
    return {"logo": data_url}

# ==================== SEED ====================

@api_router.post("/seed")
async def seed_data(user=Depends(get_user)):
    from seed_data import generate_seed_data
    await generate_seed_data(db, user["restaurant_id"])
    return {"status": "Seed data created successfully"}

# ==================== APP SETUP ====================

app.include_router(api_router)
app.add_middleware(CORSMiddleware, allow_credentials=True, allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','), allow_methods=["*"], allow_headers=["*"])

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
